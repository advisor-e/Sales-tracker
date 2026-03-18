from pathlib import Path
from collections import Counter
from difflib import SequenceMatcher
import html
from io import BytesIO
import json
import os
import re
import time
import tomllib
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode

try:
    from streamlit.runtime.scriptrunner import get_script_run_ctx
except Exception:
    get_script_run_ctx = None


def has_streamlit_context() -> bool:
    return bool(get_script_run_ctx and get_script_run_ctx() is not None)


def cache_data(*args, **kwargs):
    def decorator(func):
        if not has_streamlit_context():
            return func
        return st.cache_data(*args, **kwargs)(func)

    return decorator

BASE_DIR = Path(__file__).resolve().parent.parent
APP_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = BASE_DIR / "Get.1a.Sales Tracker.polished.xlsx"
UI_STATE_PATH = APP_DIR / "ui_state.json"
APP_CONFIG_PATH = APP_DIR / "app_config.json"
BACKUP_DIR = APP_DIR / "backups"
NAV_ITEMS = ["Home", "Dashboard", "Pipeline", "Team", "COI", "Blog"]
NEW_ROW_SENTINEL = 1_000_000
DROPDOWN_OPTIONS = {
    "Pipeline": {
        "Prospect Status": ["Active", "Await Research", "Completed", "Dead", "On Hold"],
        "Relationship Type": ["Existing Client", "New Prospect"],
        "Prospect Source": ["Social Media", "Web Enquiry", "Walk-In", "Phone-In", "Referral", "Cold Target", "Networking", "Pers' Relations"],
        "Approach Style": ["Direct Contact", "Pre Approach - Single", "Pre Approach - Sequence", "Pre Approach Gift", "Group Positioning", "Quiz Link Sent"],
        "Secure Meeting": ["No", "Yes"],
        "Quiz Completed": ["No", "Yes"],
        "Sales Style": ["Campaign", "Total Needs"],
        "Meeting Theme": ["Lite Funda' Sales", "Formal Risk Mgt", "Planning Review", "Education Session", "No Script", "Issue Response", "Full Discovery", "Growth Curve", "Other 2"],
        "Proposal Sent": ["No", "Yes"],
        "Job Secured": ["No", "Yes"],
    },
    "COI": {
        "Could We": ["No", "Yes"],
        "How Would We": ["No", "Yes"],
        "Will We": ["No", "Yes"],
        "Test/ Review": ["No", "Yes"],
    },
}

PIPELINE_IMPORT_ALIASES = {
    "prospectname": "Prospect Name",
    "clientname": "Prospect Name",
    "name": "Prospect Name",
    "businessname": "Business Name",
    "company": "Business Name",
    "companyname": "Business Name",
    "partner": "Partner",
    "leadstaffclientmanager": "Lead Staff (Client Manager)",
    "leadstaff": "Lead Staff (Client Manager)",
    "clientmanager": "Lead Staff (Client Manager)",
    "teammember": "Lead Staff (Client Manager)",
    "prospectstatus": "Prospect Status",
    "status": "Prospect Status",
    "approachdate": "Approach Date",
    "approachstyle": "Approach Style",
    "securemeeting": "Secure Meeting",
    "proposalsent": "Proposal Sent",
    "proposalrequested": "Proposal Sent",
    "proposalvalue": "Proposal Value",
    "jobsecured": "Job Secured",
    "jobsecuredvalue": "Job Secured Value",
    "commentsnextmove": "Comments / Next Move",
    "comments": "Comments / Next Move",
    "nextmove": "Comments / Next Move",
    "email": "Email",
    "contactphone": "Contact Phone #",
    "contactphone#": "Contact Phone #",
    "industry": "Industry",
    "relationshiptype": "Relationship Type",
    "prospectsource": "Prospect Source",
    "coiinvolved": "COI Involved",
}

LISTS_RANGE_MAP = {
    "Pipeline": {
        "Prospect Status": (148, 152),
        "Relationship Type": (5, 6),
        "Prospect Source": (20, 28),
        "Approach Style": (10, 16),
        "Secure Meeting": (144, 145),
        "Quiz Completed": (144, 145),
        "Sales Style": (140, 141),
        "Meeting Theme": (31, 39),
        "Proposal Sent": (144, 145),
        "Job Secured": (144, 145),
    },
    "COI": {
        "Could We": (144, 145),
        "How Would We": (144, 145),
        "Will We": (144, 145),
        "Test/ Review": (144, 145),
    },
}

DEFAULT_NEW_ROW_VALUES = {
    "Pipeline": {
        "Prospect Status": "Active",
        "Relationship Type": "New Prospect",
        "Secure Meeting": "No",
        "Quiz Completed": "No",
        "Sales Style": "Campaign",
        "Proposal Sent": "No",
        "Job Secured": "No",
        "COI Involved": "N/A",
    },
    "COI": {
        "Could We": "No",
        "How Would We": "No",
        "Will We": "No",
        "Test/ Review": "No",
    },
}
HEADER_MARKERS = {
    "Sales Activity": [
        "Prospect Name",
        "Business Name",
        "Lead Staff (Client Manager)",
        "Prospect Status",
        "Approach Date",
    ],
    "Team Report": [
        "Prospects",
        "Approaches Made",
        "Secure 1st Meeting",
        "Proposal Requested",
        "Total Proposal Value",
    ],
    "COI Development": [
        "COI Name",
        "Email",
        "Entity",
        "Lead Relationship Partner",
        "Total Referrals",
    ],
}
PAGE_THEMES = {
    "Dashboard": {"start": "#ffd166", "end": "#ff9f1c", "text": "#4a2a00"},
    "Pipeline": {"start": "#ff6b6b", "end": "#f72585", "text": "#fff7fb"},
    "Team": {"start": "#4cc9f0", "end": "#4361ee", "text": "#f7fbff"},
    "COI": {"start": "#9b5de5", "end": "#6a4c93", "text": "#fbf7ff"},
}
TABLE_CONFIG = {
    "Pipeline": [
        {"field": "Prospect Name", "header": "Prospect Name"},
        {"field": "Business Name", "header": "Business Name"},
        {"field": "Partner", "header": "Partner"},
        {"field": "Lead Staff (Client Manager)", "header": "Team Member"},
        {"field": "Prospect Status", "header": "Prospect Status"},
        {"field": "Approach Date", "header": "Approach Date"},
        {"field": "Approach Style", "header": "Approach Style"},
        {"field": "Secure Meeting", "header": "Secure Meeting"},
        {"field": "Proposal Sent", "header": "Proposal Requested"},
        {"field": "Proposal Value", "header": "Proposal Value"},
        {"field": "Job Secured", "header": "Job Secured"},
        {"field": "Job Secured Value", "header": "Job Secured Value"},
        {"field": "Comments / Next Move", "header": "Comments / Next Move"},
    ],
    "Team": [
        {"field": "Team Member", "header": "Team Member"},
        {"field": "Prospects", "header": "Prospects"},
        {"field": "Approaches Made", "header": "Approaches Made"},
        {"field": "Secure 1st Meeting", "header": "Secure 1st Meeting"},
        {"field": "Proposal Requested", "header": "Proposal Requested"},
        {"field": "Total Proposal Value", "header": "Total Proposal Value"},
        {"field": "Number of Engagements Secured", "header": "Engagements Secured"},
        {"field": "Total Value Work Secured", "header": "Work Secured Value"},
        {"field": "Avg Approach Conversion %", "header": "Avg Approach Conversion %"},
        {"field": "Avg Proposal Value", "header": "Avg Proposal Value"},
        {"field": "Avg Secured Conversion %", "header": "Avg Secured Conversion %"},
        {"field": "Active", "header": "Active"},
        {"field": "Await Research", "header": "Await Research"},
        {"field": "Completed", "header": "Completed"},
        {"field": "Dead", "header": "Dead"},
        {"field": "On Hold", "header": "On Hold"},
    ],
    "COI": [
        {"field": "COI Name", "header": "COI Name"},
        {"field": "Email", "header": "Email"},
        {"field": "Cell #", "header": "Cell #"},
        {"field": "Entity", "header": "Entity"},
        {"field": "Position", "header": "Position"},
        {"field": "Industry", "header": "Industry"},
        {"field": "Lead Relationship Partner", "header": "Lead Relationship Partner"},
        {"field": "Relationship Support", "header": "Relationship Support"},
        {"field": "Could We", "header": "Could We"},
        {"field": "How Would We", "header": "How Would We"},
        {"field": "Will We", "header": "Will We"},
        {"field": "Test/ Review", "header": "Test / Review"},
        {"field": "Total Referrals", "header": "Total Referrals"},
        {"field": "Total Converted", "header": "Total Converted"},
        {"field": "Fee Value", "header": "Fee Value"},
    ],
}

GRID_ROW_ID_JS = JsCode(
    """
    function(params) {
        if (params?.node?.rowPinned) {
            return "__pinned_top__";
        }
        const data = params?.data || {};
        const explicitId = data.__pandas_index ?? data.__row_index__;
        if (explicitId !== undefined && explicitId !== null && explicitId !== "") {
            return String(explicitId);
        }
        const rowIndex = params?.node?.rowIndex;
        return rowIndex === undefined || rowIndex === null ? "__row__unknown" : `__row__${rowIndex}`;
    }
    """
)

GRID_SHOULD_RETURN_JS = JsCode(
    """
    function should_return({streamlitRerunEventTriggerName, eventData}) {
        if (streamlitRerunEventTriggerName === "columnResized") {
            const source = eventData?.source;
            return Boolean(eventData?.finished) && !["api", "autosizeColumns", "sizeColumnsToFit"].includes(source);
        }
        if (streamlitRerunEventTriggerName === "columnMoved") {
            return Boolean(eventData?.finished);
        }
        return true;
    }
    """
)

GRID_ON_READY_JS = JsCode(
    """
    function(event) {
        const globalWindow = window;
        if (globalWindow.__salesTrackerWarnFilterInstalled) {
            return;
        }

        const blockedSnippets = [
            "rowNode.isFullWidthCell() has been deprecated",
            "server_sync_strategy is 'client_wins' and Data was edited on Grid",
            "getRowId was not set. Auto Rows hashes will be used as row ids.",
            "api.forEachDetailGridInfo",
            "MasterDetailModule"
        ];

        const originalWarn = console.warn.bind(console);
        console.warn = function(...args) {
            const message = args.map((value) => {
                if (typeof value === "string") {
                    return value;
                }
                try {
                    return JSON.stringify(value);
                } catch (error) {
                    return String(value);
                }
            }).join(" ");

            if (blockedSnippets.some((snippet) => message.includes(snippet))) {
                return;
            }

            originalWarn(...args);
        };

        globalWindow.__salesTrackerWarnFilterInstalled = true;
    }
    """
)


def apply_theme() -> None:
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Open+Sans:wght@400;600;700;800&display=swap');

        html, body, [class*="st-"], [class*="css"], .stApp, .stMarkdown, .stButton > button,
        input, textarea, select, label, p, div, span, h1, h2, h3, h4, h5, h6 {
            font-family: 'Open Sans', sans-serif !important;
        }

        :root {
            --navy: #083d77;
            --sky: #12b5ea;
            --mint: #7ee081;
            --sun: #ffd166;
            --coral: #ff6b6b;
            --paper: #fff8ef;
            --ink: #10324a;
        }

        .stApp {
            background:
                radial-gradient(circle at top left, rgba(255, 209, 102, 0.42), transparent 24%),
                radial-gradient(circle at top right, rgba(18, 181, 234, 0.28), transparent 22%),
                linear-gradient(180deg, #fffdf8 0%, #f3fbff 100%);
        }

        .block-container {
            padding-top: 1.4rem;
            padding-bottom: 2.5rem;
        }

        .nav-shell {
            background: linear-gradient(135deg, rgba(8, 61, 119, 0.96), rgba(18, 181, 234, 0.92));
            border: 3px solid rgba(255, 209, 102, 0.95);
            border-radius: 22px;
            box-shadow: 0 16px 36px rgba(8, 61, 119, 0.18);
            color: white;
            margin-bottom: 1rem;
            padding: 0.85rem 1.1rem 1rem;
        }

        .nav-shell h2 {
            color: white;
            font-size: 1.05rem;
            letter-spacing: 0.08em;
            margin: 0;
            text-transform: uppercase;
        }

        .hero-card {
            background: linear-gradient(135deg, rgba(255, 107, 107, 0.95), rgba(255, 209, 102, 0.95));
            border-radius: 26px;
            box-shadow: 0 18px 42px rgba(255, 107, 107, 0.22);
            color: #2e2244;
            margin-bottom: 1.2rem;
            overflow: hidden;
            padding: 1.45rem 1.5rem;
        }

        .hero-kicker {
            font-size: 0.88rem;
            font-weight: 700;
            letter-spacing: 0.08em;
            margin-bottom: 0.2rem;
            text-transform: uppercase;
        }

        .hero-title {
            font-size: 2.1rem;
            font-weight: 800;
            line-height: 1.05;
            margin-bottom: 0.45rem;
        }

        .hero-copy {
            font-size: 1rem;
            line-height: 1.5;
            margin-bottom: 0;
            max-width: 52rem;
        }

        .quick-links {
            display: grid;
            gap: 0.85rem;
            grid-template-columns: repeat(auto-fit, minmax(210px, 1fr));
            margin: 0.35rem 0 1.2rem;
        }

        .quick-link {
            background: rgba(255, 255, 255, 0.84);
            border: 2px solid rgba(18, 181, 234, 0.24);
            border-radius: 20px;
            box-shadow: 0 12px 24px rgba(16, 50, 74, 0.08);
            padding: 1rem;
        }

        .quick-link strong {
            color: var(--navy);
            display: block;
            font-size: 1rem;
            margin-bottom: 0.25rem;
        }

        .quick-link span {
            color: var(--ink);
            font-size: 0.93rem;
        }

        .section-banner {
            background: linear-gradient(135deg, rgba(126, 224, 129, 0.95), rgba(18, 181, 234, 0.88));
            border-radius: 22px;
            box-shadow: 0 14px 32px rgba(18, 181, 234, 0.14);
            color: #073b4c;
            margin: 0.3rem 0 1rem;
            padding: 1rem 1.2rem;
        }

        .section-banner h3 {
            color: #073b4c;
            font-size: 1.45rem;
            margin: 0 0 0.2rem;
        }

        .section-banner p {
            margin: 0;
        }

        div[data-testid="stRadio"] label {
            font-weight: 700;
        }

        div[role="radiogroup"] {
            gap: 0.5rem;
            flex-wrap: wrap;
            row-gap: 0.55rem;
        }

        div[role="radiogroup"] label[data-baseweb="radio"] {
            background: rgba(255, 255, 255, 0.16);
            border: 2px solid rgba(255, 255, 255, 0.22);
            border-radius: 999px;
            cursor: pointer;
            padding: 0.45rem 0.9rem;
            transition: background-color 0.18s ease, border-color 0.18s ease, box-shadow 0.18s ease, transform 0.18s ease;
        }

        div[role="radiogroup"] label[data-baseweb="radio"]:hover {
            background: rgba(255, 209, 102, 0.28);
            border-color: rgba(255, 248, 239, 0.88);
            box-shadow: 0 10px 20px rgba(8, 61, 119, 0.22);
            transform: translateY(-1px);
        }

        div[role="radiogroup"] label[data-baseweb="radio"]:focus-within {
            background: rgba(126, 224, 129, 0.28);
            border-color: #ffffff;
            box-shadow: 0 0 0 3px rgba(255, 209, 102, 0.32);
        }

        div[role="radiogroup"] label[data-baseweb="radio"]:has(input:checked) {
            background: linear-gradient(135deg, #ffd166, #ff6b6b);
            border-color: #fff8ef;
            box-shadow: 0 8px 16px rgba(255, 107, 107, 0.2);
        }

        div[role="radiogroup"] label[data-baseweb="radio"]:has(input:checked):hover {
            background: linear-gradient(135deg, #ffe08a, #ff8d76);
            border-color: #ffffff;
            box-shadow: 0 12px 24px rgba(255, 107, 107, 0.28);
        }

        div[role="radiogroup"] label[data-baseweb="radio"] p {
            color: #111111;
            text-shadow: none;
        }

        div[role="radiogroup"] label[data-baseweb="radio"]:hover p,
        div[role="radiogroup"] label[data-baseweb="radio"]:focus-within p {
            color: #111111;
        }

        div[role="radiogroup"] label[data-baseweb="radio"]:has(input:checked) p {
            color: #111111;
            text-shadow: none;
        }

        div[data-testid="metric-container"] {
            background: rgba(255, 255, 255, 0.84);
            border: 2px solid rgba(255, 209, 102, 0.34);
            border-radius: 18px;
            box-shadow: 0 12px 28px rgba(16, 50, 74, 0.08);
            padding: 0.8rem 0.9rem;
        }

        div[data-testid="stTabs"] button[role="tab"] {
            border-radius: 999px;
        }

        .stButton > button[kind="primary"] {
            background: linear-gradient(135deg, #34d399, #16a34a) !important;
            border: 2px solid #0f9f43 !important;
            color: #ffffff !important;
            font-weight: 700 !important;
        }

        .stButton > button[kind="primary"]:hover {
            background: linear-gradient(135deg, #4ade80, #22c55e) !important;
            border-color: #15803d !important;
        }

        </style>
        """,
        unsafe_allow_html=True,
    )


def initialize_navigation() -> None:
    if "current_page" not in st.session_state:
        st.session_state.current_page = "Home"
    if "dropdown_options" not in st.session_state:
        st.session_state["dropdown_options"] = {
            view: {field: list(opts) for field, opts in fields.items()}
            for view, fields in DROPDOWN_OPTIONS.items()
        }
    ui_state = load_ui_state()
    for view in ("pipeline", "team", "coi"):
        state_key = f"{view}_columns_state"
        if state_key not in st.session_state:
            st.session_state[state_key] = ui_state.get(state_key)


def load_ui_state() -> dict:
    if not UI_STATE_PATH.exists():
        return {}
    try:
        return json.loads(UI_STATE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def save_ui_state(state: dict) -> None:
    try:
        UI_STATE_PATH.write_text(json.dumps(state, indent=2), encoding="utf-8")
    except OSError:
        pass


def load_app_config() -> dict:
    if not APP_CONFIG_PATH.exists():
        return {}
    try:
        return json.loads(APP_CONFIG_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def save_app_config(config_data: dict) -> None:
    try:
        APP_CONFIG_PATH.write_text(json.dumps(config_data, indent=2), encoding="utf-8")
    except OSError:
        pass


def create_workbook_backup(workbook_path: str) -> Path | None:
    source = Path(workbook_path)
    if not source.exists() or not source.is_file():
        return None

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{source.stem}_{timestamp}{source.suffix}"
    backup_path = BACKUP_DIR / backup_name
    backup_path.write_bytes(source.read_bytes())
    return backup_path


def set_persistent_columns_state(view_name: str, columns_state: list[dict] | None) -> None:
    state_key = f"{view_name.lower()}_columns_state"
    ui_state = load_ui_state()
    if columns_state:
        ui_state[state_key] = columns_state
    else:
        ui_state.pop(state_key, None)
    save_ui_state(ui_state)


def render_navigation(is_manager: bool = False) -> str:
    nav_items = list(NAV_ITEMS)
    if is_manager and "Lists" not in nav_items:
        nav_items.insert(1, "Lists")
    dashboard_idx = nav_items.index("Dashboard") + 1 if "Dashboard" in nav_items else -1
    pipeline_idx = nav_items.index("Pipeline") + 1 if "Pipeline" in nav_items else -1
    team_idx = nav_items.index("Team") + 1 if "Team" in nav_items else -1
    coi_idx = nav_items.index("COI") + 1 if "COI" in nav_items else -1
    blog_idx = nav_items.index("Blog") + 1 if "Blog" in nav_items else -1
    # Guard against stale page value when manager signs out
    if st.session_state.get("current_page") not in nav_items:
        st.session_state["current_page"] = "Home"
    st.markdown('<div class="nav-shell"></div>', unsafe_allow_html=True)
    selected_page = st.radio(
        "Go to",
        nav_items,
        key="current_page",
        horizontal=True,
        label_visibility="collapsed",
    )

    nav_color_specs = [
        (dashboard_idx, "#ff9f1c", "#f97316", "#fed7aa", "#fb923c", "rgba(249, 115, 22, 0.28)"),
        (pipeline_idx, "#f472b6", "#db2777", "#fbcfe8", "#f472b6", "rgba(219, 39, 119, 0.28)"),
        (team_idx, "#38bdf8", "#2563eb", "#bfdbfe", "#60a5fa", "rgba(37, 99, 235, 0.28)"),
        (coi_idx, "#a78bfa", "#7c3aed", "#ddd6fe", "#a78bfa", "rgba(124, 58, 237, 0.28)"),
        (blog_idx, "#22c55e", "#15803d", "#86efac", "#22c55e", "rgba(21, 128, 61, 0.28)"),
    ]
    nav_scope = '.st-key-current_page div[role="radiogroup"] label[data-baseweb="radio"]'
    fallback_scope = 'div[data-testid="stRadio"]:has(input[aria-label="Go to"]) div[role="radiogroup"] label[data-baseweb="radio"]'
    nav_rules: list[str] = []
    for idx, base_start, base_end, checked_start, checked_end, shadow in nav_color_specs:
        if idx <= 0:
            continue
        nav_rules.append(
            f"""
            {nav_scope}:nth-of-type({idx}),
            {fallback_scope}:nth-of-type({idx}) {{
                background: linear-gradient(135deg, {base_start}, {base_end}) !important;
                border-color: rgba(248, 250, 252, 0.92) !important;
            }}
            {nav_scope}:nth-of-type({idx}):hover,
            {fallback_scope}:nth-of-type({idx}):hover {{
                background: linear-gradient(135deg, {checked_start}, {base_end}) !important;
                border-color: rgba(255, 255, 255, 1) !important;
            }}
            {nav_scope}:nth-of-type({idx}):has(input:checked),
            {fallback_scope}:nth-of-type({idx}):has(input:checked) {{
                background: linear-gradient(135deg, {checked_start}, {checked_end}) !important;
                border-color: rgba(255, 255, 255, 0.98) !important;
                box-shadow: 0 10px 22px {shadow} !important;
            }}
            {nav_scope}:nth-of-type({idx}):has(input:checked):hover,
            {fallback_scope}:nth-of-type({idx}):has(input:checked):hover {{
                background: linear-gradient(135deg, {checked_start}, {checked_end}) !important;
                border-color: rgba(255, 255, 255, 1) !important;
            }}
            """
        )

    st.markdown(
        f"""
        <style>
        {''.join(nav_rules)}
        </style>
        """,
        unsafe_allow_html=True,
    )
    return selected_page


def render_home_page(filtered_sales: pd.DataFrame, team: pd.DataFrame, coi: pd.DataFrame) -> None:
    secured_value = float(filtered_sales.get("Job Secured Value", pd.Series(dtype=float)).sum())
    active_staff = int(filtered_sales.get("Lead Staff (Client Manager)", pd.Series(dtype=str)).dropna().astype(str).str.strip().ne("").sum())
    prospect_count = int(len(filtered_sales))
    coi_count = int(len(coi.index))

    st.markdown(
        """
        <div class="hero-card">
            <div class="hero-kicker">Home Page</div>
            <div class="hero-title">Find your way around the sales tracker fast.</div>
            <p class="hero-copy">
                Use the bright navigation bar above to jump between the dashboard, charts, pipeline, team, and COI views.
                The Home page stays available as a clear reset point whenever you want to come back here.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Filtered prospects", f"{prospect_count:,}")
    c2.metric("Active staff entries", f"{active_staff:,}")
    c3.metric("Secured value", f"${secured_value:,.0f}")
    c4.metric("COI records", f"{coi_count:,}")

    st.markdown(
        """
        <div class="quick-links">
            <div class="quick-link"><strong>Dashboard</strong><span>Top-line KPIs for approaches, meetings, proposals, and secured work.</span></div>
            <div class="quick-link"><strong>Dashboard Charts</strong><span>See the KPI summary and the visual trends together in one place.</span></div>
            <div class="quick-link"><strong>Pipeline</strong><span>Detailed pipeline table with CSV download for filtered results.</span></div>
            <div class="quick-link"><strong>Team</strong><span>Open the team report without digging through tabs.</span></div>
            <div class="quick-link"><strong>COI</strong><span>Jump straight to referral development activity and records.</span></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    left, right = st.columns([1.3, 1])
    with left:
        st.subheader("Today at a glance")
        preview_columns = [
            "Prospect Name",
            "Business Name",
            "Lead Staff (Client Manager)",
            "Prospect Status",
            "Job Secured Value",
        ]
        existing_columns = [column for column in preview_columns if column in filtered_sales.columns]
        st.dataframe(filtered_sales[existing_columns].head(8), use_container_width=True, height=320)

    with right:
        st.subheader("What each section gives you")
        st.info("Dashboard now includes both headline numbers and charts.")
        st.success("Visual trends sit directly under the KPI summary.")
        st.warning("Pipeline is the working view for exporting and follow-up.")
        st.caption(f"Team rows loaded: {len(team.index):,} | COI rows loaded: {len(coi.index):,}")


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def canonicalize_column_name(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def next_new_row_start(df: pd.DataFrame) -> int:
    if df.empty:
        return NEW_ROW_SENTINEL
    numeric_index = pd.to_numeric(pd.Index(df.index), errors="coerce")
    sentinel_values = [int(v) for v in numeric_index if pd.notna(v) and int(v) >= NEW_ROW_SENTINEL]
    return (max(sentinel_values) + 1) if sentinel_values else NEW_ROW_SENTINEL


def build_pipeline_base_df(source_df: pd.DataFrame, pipeline_columns: list[str]) -> pd.DataFrame:
    result = source_df.copy()
    for column in pipeline_columns:
        if column not in result.columns:
            result[column] = ""
    ordered_columns = [column for column in pipeline_columns if column in result.columns]
    return result[ordered_columns].copy()


def load_import_source(uploaded_file) -> pd.DataFrame:
    suffix = Path(uploaded_file.name).suffix.lower()
    uploaded_file.seek(0)
    if suffix == ".csv":
        return pd.read_csv(uploaded_file)
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(uploaded_file)
    raise ValueError("Only .csv, .xlsx, .xlsm, and .xls files are supported.")


def get_pipeline_import_mapping(
    imported_columns: list[str],
    pipeline_columns: list[str],
    column_overrides: dict[str, str] | None = None,
) -> tuple[pd.DataFrame, list[str], list[str]]:
    target_by_key = {canonicalize_column_name(column): column for column in pipeline_columns}
    for alias_key, target_name in PIPELINE_IMPORT_ALIASES.items():
        if target_name in pipeline_columns:
            target_by_key.setdefault(alias_key, target_name)

    column_overrides = column_overrides or {}

    preview_rows: list[dict[str, str]] = []
    matched_columns: list[str] = []
    unmapped_columns: list[str] = []

    for column in imported_columns:
        override_value = (column_overrides.get(str(column)) or "").strip()
        if override_value == "Ignore":
            target_name = ""
        elif override_value:
            target_name = override_value
        else:
            target_name = target_by_key.get(canonicalize_column_name(column), "")

        if target_name:
            matched_columns.append(target_name)
            preview_rows.append(
                {
                    "Incoming Column": str(column),
                    "Pipeline Field": target_name,
                    "Status": "Matched",
                }
            )
        else:
            unmapped_columns.append(str(column))
            preview_rows.append(
                {
                    "Incoming Column": str(column),
                    "Pipeline Field": "",
                    "Status": "Ignored",
                }
            )

    preview_df = pd.DataFrame(preview_rows)
    return preview_df, list(dict.fromkeys(matched_columns)), unmapped_columns


def build_pipeline_import_template(
    pipeline_columns: list[str],
    dropdown_options: dict[str, dict[str, list[str]]],
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Pipeline Import Template"
    header_fill = PatternFill(fill_type="solid", fgColor="0F4C81")
    header_font = Font(color="FFFFFF", bold=True)
    sample_fill = PatternFill(fill_type="solid", fgColor="E0F2FE")

    sample_row = {
        "Prospect Name": "Jane Example",
        "Business Name": "Example Pty Ltd",
        "Partner": "Partner Name",
        "Lead Staff (Client Manager)": "Advisor Name",
        "Prospect Status": "Active",
        "Approach Date": "2026-03-18",
        "Approach Style": "Direct Contact",
        "Secure Meeting": "No",
        "Proposal Sent": "No",
        "Proposal Value": "0",
        "Job Secured": "No",
        "Job Secured Value": "0",
        "Comments / Next Move": "Replace this sample row with your real data.",
        "Relationship Type": "New Prospect",
        "Prospect Source": "Referral",
        "COI Involved": "N/A",
    }

    for col_idx, header in enumerate(pipeline_columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max(16, min(len(header) + 4, 28))

        sample_cell = worksheet.cell(row=2, column=col_idx, value=sample_row.get(header, ""))
        sample_cell.fill = sample_fill

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = True

    instructions_sheet = workbook.create_sheet("Instructions", 0)
    instructions_sheet["A1"] = "Pipeline Import Instructions"
    instructions_sheet["A1"].font = Font(bold=True, size=14)
    instruction_lines = [
        "1. Use the 'Pipeline Import Template' sheet for your data.",
        "2. Keep the header row unchanged so the app can match fields correctly.",
        "3. Row 2 is a sample row only. Delete or replace it before import if not needed.",
        "4. Dropdown cells already include the same allowed values used in the app.",
        "5. Dates are best entered as YYYY-MM-DD.",
        "6. Save the file as .xlsx and upload it from the Pipeline page.",
        "7. You can still review and edit imported rows in the grid before saving to the workbook.",
    ]
    for row_idx, line in enumerate(instruction_lines, start=3):
        instructions_sheet.cell(row=row_idx, column=1, value=line)
    instructions_sheet.column_dimensions["A"].width = 100

    validation_sheet = workbook.create_sheet("Validation Lists")
    validation_sheet.sheet_state = "hidden"
    validation_fields = dropdown_options.get("Pipeline", {})

    validation_col = 1
    for field in pipeline_columns:
        values = validation_fields.get(field)
        if not values:
            continue

        validation_sheet.cell(row=1, column=validation_col, value=field)
        for row_idx, value in enumerate(values, start=2):
            validation_sheet.cell(row=row_idx, column=validation_col, value=value)

        target_col_letter = get_column_letter(pipeline_columns.index(field) + 1)
        list_col_letter = get_column_letter(validation_col)
        formula = f"'Validation Lists'!${list_col_letter}$2:${list_col_letter}${len(values) + 1}"
        data_validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        data_validation.prompt = f"Choose a value for {field}"
        data_validation.promptTitle = field
        worksheet.add_data_validation(data_validation)
        data_validation.add(f"{target_col_letter}2:{target_col_letter}500")
        validation_col += 1

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output.getvalue()


def parse_pipeline_import(
    uploaded_file,
    pipeline_columns: list[str],
    column_overrides: dict[str, str] | None = None,
) -> tuple[pd.DataFrame, list[str], list[str]]:
    imported_df = load_import_source(uploaded_file)
    imported_df = normalize_columns(imported_df)

    preview_df, matched_columns, unmapped_columns = get_pipeline_import_mapping(
        imported_df.columns.tolist(),
        pipeline_columns,
        column_overrides=column_overrides,
    )

    rename_map = {
        row["Incoming Column"]: row["Pipeline Field"]
        for row in preview_df.to_dict(orient="records")
        if row.get("Status") == "Matched" and row.get("Pipeline Field")
    }

    imported_df = imported_df.rename(columns=rename_map)
    if not matched_columns:
        raise ValueError("No import columns matched the Pipeline fields.")

    pipeline_import_df = pd.DataFrame(columns=pipeline_columns)
    for column in pipeline_columns:
        if column in imported_df.columns:
            pipeline_import_df[column] = imported_df[column]
        else:
            pipeline_import_df[column] = ""

    pipeline_import_df = pipeline_import_df.fillna("")
    if "Approach Date" in pipeline_import_df.columns:
        parsed_dates = pd.to_datetime(pipeline_import_df["Approach Date"], errors="coerce")
        pipeline_import_df["Approach Date"] = parsed_dates.where(parsed_dates.notna(), pipeline_import_df["Approach Date"])

    pipeline_import_df = pipeline_import_df.reset_index(drop=True)
    return pipeline_import_df, matched_columns, unmapped_columns


def append_imported_rows(df: pd.DataFrame, imported_rows: pd.DataFrame | None) -> pd.DataFrame:
    if imported_rows is None or imported_rows.empty:
        return df

    result = df.copy()
    imported = imported_rows.copy()
    for column in result.columns:
        if column not in imported.columns:
            imported[column] = ""
    imported = imported[result.columns]

    start_index = next_new_row_start(result)
    imported.index = range(start_index, start_index + len(imported))
    return pd.concat([result, imported], axis=0)


def find_header_row(workbook_path: str, sheet_name: str, markers: list[str]) -> int:
    preview = pd.read_excel(workbook_path, sheet_name=sheet_name, header=None, nrows=15, engine="openpyxl")
    best_row = 0
    best_score = -1

    for row_index in range(len(preview.index)):
        row_values = {str(value).strip() for value in preview.iloc[row_index].tolist() if pd.notna(value)}
        score = sum(1 for marker in markers if marker in row_values)
        if score > best_score:
            best_score = score
            best_row = row_index

    return best_row


def load_sheet(workbook_path: str, sheet_name: str, markers: list[str] | None = None) -> pd.DataFrame:
    header_row = find_header_row(workbook_path, sheet_name, markers) if markers else 0
    df = pd.read_excel(workbook_path, sheet_name=sheet_name, header=header_row, engine="openpyxl")
    df = normalize_columns(df)
    df = df.dropna(axis=1, how="all")
    data_start_excel_row = header_row + 2
    df["__excel_row__"] = range(data_start_excel_row, data_start_excel_row + len(df))

    if sheet_name == "Sales Activity":
        df = df[df.get("Prospect Name", pd.Series(dtype=object)).notna()].copy()
    elif sheet_name == "Team Report":
        staff_column = df.columns[3] if len(df.columns) > 3 else df.columns[0]
        df = df.rename(columns={staff_column: "Team Member"})
        team_member = df["Team Member"].astype(str).str.strip()
        df = df[df["Team Member"].notna() & ~team_member.isin(["", "0", "0.0", "nan"])].copy()
    elif sheet_name == "COI Development":
        df = df[df.get("COI Name", pd.Series(dtype=object)).notna()].copy()

    return df.reset_index(drop=True)


def build_team_report(workbook_path: str, sales_df: pd.DataFrame) -> pd.DataFrame:
    """Build the Team Report by aggregating Sales Activity data.

    Team member names are read from Lists!B82:B106 where they are stored as
    literal text values (not formulas). All statistics are computed from the
    already-coerced sales_df so we bypass the formula-evaluation limitation.
    """
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    ws = wb["Lists"]
    names: list[str] = []
    for row in ws.iter_rows(min_row=82, max_row=106, min_col=2, max_col=2, values_only=True):
        val = row[0]
        if val is not None and isinstance(val, str) and not val.startswith("=") and val.strip():
            names.append(val.strip())
    wb.close()

    # Fallback: derive names from Sales Activity if Lists had none
    if not names:
        staff_col = "Lead Staff (Client Manager)"
        if staff_col in sales_df.columns:
            names = sorted(
                sales_df[staff_col].dropna().astype(str).str.strip().unique().tolist()
            )

    staff_col = "Lead Staff (Client Manager)"
    status_col = "Prospect Status"
    rows: list[dict] = []
    for name in names:
        if staff_col in sales_df.columns:
            mdf = sales_df[sales_df[staff_col].astype(str).str.strip() == name]
        else:
            mdf = pd.DataFrame()

        prospects = len(mdf)
        approaches = int(
            mdf.get("Approach Style", pd.Series(dtype=str))
            .fillna("").astype(str).str.strip().ne("").sum()
        )
        meetings = int(
            (mdf.get("Secure Meeting", pd.Series(dtype=str))
             .fillna("").astype(str).str.lower() == "yes").sum()
        )
        proposals = int(
            (mdf.get("Proposal Sent", pd.Series(dtype=str))
             .fillna("").astype(str).str.lower() == "yes").sum()
        )
        total_proposal_val = float(
            mdf.get("Proposal Value", pd.Series(dtype=float)).sum()
        )
        secured = int(
            (mdf.get("Job Secured", pd.Series(dtype=str))
             .fillna("").astype(str).str.lower() == "yes").sum()
        )
        secured_val = float(
            mdf.get("Job Secured Value", pd.Series(dtype=float)).sum()
        )
        avg_approach = round(approaches / prospects, 4) if prospects > 0 else 0.0
        avg_proposal_val = round(total_proposal_val / proposals, 2) if proposals > 0 else 0.0
        avg_secured = round(secured / proposals, 4) if proposals > 0 else 0.0
        statuses = (
            mdf.get(status_col, pd.Series(dtype=str)).fillna("").astype(str).str.strip()
        )
        rows.append({
            "Team Member": name,
            "Prospects": prospects,
            "Approaches Made": approaches,
            "Secure 1st Meeting": meetings,
            "Proposal Requested": proposals,
            "Total Proposal Value": total_proposal_val,
            "Number of Engagements Secured": secured,
            "Total Value Work Secured": secured_val,
            "Avg Approach Conversion %": avg_approach,
            "Avg Proposal Value": avg_proposal_val,
            "Avg Secured Conversion %": avg_secured,
            "Active": int((statuses == "Active").sum()),
            "Await Research": int((statuses == "Await Research").sum()),
            "Completed": int((statuses == "Completed").sum()),
            "Dead": int((statuses == "Dead").sum()),
            "On Hold": int((statuses == "On Hold").sum()),
        })
    return pd.DataFrame(rows)


def coerce_currency(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace(" ", "", regex=False)
        .replace({"": None, "-": None, "None": None, "nan": None})
    )
    return pd.to_numeric(cleaned, errors="coerce").fillna(0)


def resolve_lists_reference(value: object, lists_ws) -> object:
    if not isinstance(value, str):
        return value
    match = re.fullmatch(r"=Lists!\$?([A-Z]+)\$?(\d+)", value.strip())
    if not match:
        return value
    ref = f"{match.group(1)}{match.group(2)}"
    return lists_ws[ref].value


def split_formula_args(arg_text: str) -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    depth = 0
    in_quotes = False
    i = 0
    while i < len(arg_text):
        ch = arg_text[i]
        if ch == '"':
            in_quotes = not in_quotes
            current.append(ch)
        elif ch == ',' and not in_quotes and depth == 0:
            parts.append("".join(current).strip())
            current = []
        else:
            if ch == '(' and not in_quotes:
                depth += 1
            elif ch == ')' and not in_quotes and depth > 0:
                depth -= 1
            current.append(ch)
        i += 1
    if current:
        parts.append("".join(current).strip())
    return parts


def to_number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100.0
        except ValueError:
            return 0.0
    if text.startswith("$"):
        text = text[1:]
    try:
        return float(text)
    except ValueError:
        return 0.0


def normalize_for_compare(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def col_to_index(col: str) -> int:
    index = 0
    for ch in col:
        index = (index * 26) + (ord(ch) - ord("A") + 1)
    return index


def parse_cell_ref(ref: str) -> tuple[int, int] | None:
    m = re.fullmatch(r"\$?([A-Z]+)\$?(\d+)", ref.strip())
    if not m:
        return None
    return int(m.group(2)), col_to_index(m.group(1))


def iter_sales_range_values(range_ref: str, sales_ws) -> list[object]:
    m = re.fullmatch(
        r"'Sales Activity'!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)",
        range_ref.strip(),
    )
    if not m:
        return []
    col1, row1, col2, row2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    c1 = col_to_index(col1)
    c2 = col_to_index(col2)
    values: list[object] = []
    for row in range(min(row1, row2), max(row1, row2) + 1):
        for col in range(min(c1, c2), max(c1, c2) + 1):
            values.append(sales_ws.cell(row=row, column=col).value)
    return values


@cache_data(show_spinner=False)
def get_sheet_columns_by_excel_range(
    workbook_path: str,
    sheet_name: str,
    start_col: str,
    end_col: str,
    markers: list[str],
) -> list[str]:
    header_row_index = find_header_row(workbook_path, sheet_name, markers)
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    ws = wb[sheet_name]
    header_row_excel = header_row_index + 1
    start_idx = col_to_index(start_col)
    end_idx = col_to_index(end_col)

    columns: list[str] = []
    seen: set[str] = set()
    for col_idx in range(start_idx, end_idx + 1):
        value = ws.cell(row=header_row_excel, column=col_idx).value
        name = str(value).strip() if value is not None else ""
        if name and not name.lower().startswith("unnamed") and name not in seen:
            columns.append(name)
            seen.add(name)

    wb.close()
    return columns


def load_stats_to_date_detail(workbook_path: str) -> pd.DataFrame:
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    stats_ws = wb["Stats to Date"]
    lists_ws = wb["Lists"]
    sales_ws = wb["Sales Activity"]

    stats_cache: dict[tuple[int, int], object] = {}
    sales_range_cache: dict[str, list[object]] = {}

    def get_sales_range(range_ref: str) -> list[object]:
        key = range_ref.strip()
        if key not in sales_range_cache:
            sales_range_cache[key] = iter_sales_range_values(key, sales_ws)
        return sales_range_cache[key]

    def eval_token(token: str) -> object:
        tok = token.strip()
        if tok.startswith('"') and tok.endswith('"'):
            return tok[1:-1]
        if tok.startswith("="):
            return eval_formula(tok)
        div_token = re.fullmatch(r"(\$?[A-Z]+\$?\d+)\/(\$?[A-Z]+\$?\d+)", tok)
        if div_token:
            numerator = to_number(eval_token(div_token.group(1)))
            denominator = to_number(eval_token(div_token.group(2)))
            return 0.0 if denominator == 0 else numerator / denominator
        parsed_ref = parse_cell_ref(tok)
        if parsed_ref is not None:
            return eval_stats_cell(parsed_ref[0], parsed_ref[1])
        try:
            if "." in tok:
                return float(tok)
            return int(tok)
        except ValueError:
            return tok

    def eval_formula(formula: str) -> object:
        f = formula.strip()
        if not f.startswith("="):
            return f

        lists_match = re.fullmatch(r"=Lists!\$?([A-Z]+)\$?(\d+)", f)
        if lists_match:
            return lists_ws[f"{lists_match.group(1)}{lists_match.group(2)}"].value

        fn_match = re.fullmatch(r"=([A-Z]+)\((.*)\)", f)
        if fn_match:
            fn = fn_match.group(1)
            args = split_formula_args(fn_match.group(2))

            if fn == "COUNTIF" and len(args) == 2:
                vals = get_sales_range(args[0])
                crit = normalize_for_compare(eval_token(args[1]))
                return sum(1 for v in vals if normalize_for_compare(v) == crit)

            if fn == "COUNTIFS" and len(args) >= 2 and len(args) % 2 == 0:
                pairs = [(args[i], eval_token(args[i + 1])) for i in range(0, len(args), 2)]
                value_lists = [get_sales_range(range_ref) for range_ref, _ in pairs]
                if not value_lists:
                    return 0
                n = min(len(v) for v in value_lists)
                count = 0
                for idx in range(n):
                    if all(
                        normalize_for_compare(value_lists[p_idx][idx]) == normalize_for_compare(criteria)
                        for p_idx, (_, criteria) in enumerate(pairs)
                    ):
                        count += 1
                return count

            if fn == "SUMIFS" and len(args) >= 3 and len(args) % 2 == 1:
                sum_values = get_sales_range(args[0])
                pairs = [(args[i], eval_token(args[i + 1])) for i in range(1, len(args), 2)]
                criteria_values = [get_sales_range(range_ref) for range_ref, _ in pairs]
                n = min([len(sum_values)] + [len(v) for v in criteria_values])
                total = 0.0
                for idx in range(n):
                    if all(
                        normalize_for_compare(criteria_values[p_idx][idx]) == normalize_for_compare(criteria)
                        for p_idx, (_, criteria) in enumerate(pairs)
                    ):
                        total += to_number(sum_values[idx])
                return total

            if fn == "SUM" and len(args) == 1:
                rm = re.fullmatch(r"\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", args[0].strip())
                if rm:
                    r1, c1 = int(rm.group(2)), col_to_index(rm.group(1))
                    r2, c2 = int(rm.group(4)), col_to_index(rm.group(3))
                    total = 0.0
                    for rr in range(min(r1, r2), max(r1, r2) + 1):
                        for cc in range(min(c1, c2), max(c1, c2) + 1):
                            total += to_number(eval_stats_cell(rr, cc))
                    return total

            if fn == "IF" and len(args) == 3:
                condition = args[0].strip()
                cm = re.fullmatch(r"(.+?)=(.+)", condition)
                if cm:
                    left = eval_token(cm.group(1).strip())
                    right = eval_token(cm.group(2).strip())
                    truthy = normalize_for_compare(left) == normalize_for_compare(right)
                    return eval_token(args[1].strip()) if truthy else eval_token(args[2].strip())

        div_match = re.fullmatch(r"=(\$?[A-Z]+\$?\d+)\/(\$?[A-Z]+\$?\d+)", f)
        if div_match:
            numerator = to_number(eval_token(div_match.group(1)))
            denominator = to_number(eval_token(div_match.group(2)))
            return 0.0 if denominator == 0 else numerator / denominator

        # Fallback: keep text labels where possible; mark unsupported formulas clearly.
        resolved = resolve_lists_reference(f, lists_ws)
        return "Unsupported formula" if isinstance(resolved, str) and resolved.startswith("=") else resolved

    def eval_stats_cell(row_num: int, col_num: int) -> object:
        key = (row_num, col_num)
        if key in stats_cache:
            return stats_cache[key]

        raw = stats_ws.cell(row=row_num, column=col_num).value
        if isinstance(raw, str) and raw.startswith("="):
            val = eval_formula(raw)
        else:
            val = raw
        if val is None:
            val = ""
        stats_cache[key] = val
        return val

    rows: list[dict[str, object]] = []
    for row_num in range(2, 67):
        row_payload = {"Row": row_num}
        for col_num, col_name in zip(range(2, 8), ["B", "C", "D", "E", "F", "G"]):
            value = eval_stats_cell(row_num, col_num)
            if isinstance(value, str):
                row_payload[col_name] = value.strip()
            elif isinstance(value, float):
                row_payload[col_name] = int(value) if value.is_integer() else round(value, 4)
            else:
                row_payload[col_name] = value

        if any(str(row_payload[c]).strip() for c in ["B", "C", "D", "E", "F", "G"]):
            rows.append(row_payload)

    wb.close()
    return pd.DataFrame(rows)


@cache_data(show_spinner=False)
def load_data(workbook_path: str) -> dict[str, pd.DataFrame]:
    sales = load_sheet(workbook_path, "Sales Activity", HEADER_MARKERS["Sales Activity"])
    stats = load_stats_to_date_detail(workbook_path)
    coi = load_sheet(workbook_path, "COI Development", HEADER_MARKERS["COI Development"])

    for col in ["Approach Date", "Date Secured", "Date Last Contact", "Meeting Date"]:
        if col in sales.columns:
            sales[col] = pd.to_datetime(sales[col], errors="coerce")

    if "Job Secured Value" in sales.columns:
        sales["Job Secured Value"] = coerce_currency(sales["Job Secured Value"])
    else:
        sales["Job Secured Value"] = 0

    if "Proposal Value" in sales.columns:
        sales["Proposal Value"] = coerce_currency(sales["Proposal Value"])
    else:
        sales["Proposal Value"] = 0

    # Build team report from live Sales Activity data (Team Report sheet is
    # formula-only and cannot be evaluated by openpyxl without Excel)
    team = build_team_report(workbook_path, sales)

    return {
        "sales": sales,
        "team": team,
        "stats": stats,
        "coi": coi,
    }


def app_header() -> None:
    st.title("Sales Tracker App")


def validate_workbook_path(path_text: str) -> tuple[bool, str | None]:
    path = Path(path_text).expanduser()
    if not path.exists():
        return False, "Workbook path does not exist."
    if not path.is_file():
        return False, "Workbook path must point to a file."
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False, "Workbook must be an .xlsx or .xlsm file."
    return True, None


def get_manager_password() -> str | None:
    app_config_password = load_app_config().get("manager_password")
    if app_config_password:
        return str(app_config_password)

    for env_name in ("SALES_TRACKER_MANAGER_PASSWORD", "MANAGER_PASSWORD"):
        env_password = os.getenv(env_name)
        if env_password:
            return env_password

    candidate_paths = [APP_DIR / ".streamlit" / "secrets.toml", Path.home() / ".streamlit" / "secrets.toml"]
    for path in candidate_paths:
        if not path.exists():
            continue
        try:
            with path.open("rb") as handle:
                secrets_data = tomllib.load(handle)
        except (OSError, tomllib.TOMLDecodeError):
            continue
        password = secrets_data.get("manager_password") or secrets_data.get("MANAGER_PASSWORD")
        if password:
            return str(password)

    return None


def reset_runtime_dropdown_options(coi_df: pd.DataFrame) -> None:
    st.session_state["dropdown_options"] = {
        view: {field: list(opts) for field, opts in fields.items()}
        for view, fields in DROPDOWN_OPTIONS.items()
    }

    if "COI Name" in coi_df.columns:
        coi_names = [
            str(v).strip()
            for v in coi_df["COI Name"].dropna().unique().tolist()
            if str(v).strip() and str(v).strip() != "nan"
        ]
    else:
        coi_names = []

    st.session_state["dropdown_options"].setdefault("Pipeline", {})["COI Involved"] = ["N/A"] + sorted(coi_names)


def section_banner(title: str, copy: str) -> None:
    theme = PAGE_THEMES.get(title, {"start": "rgba(126, 224, 129, 0.95)", "end": "rgba(18, 181, 234, 0.88)", "text": "#073b4c"})
    st.markdown(
        f"""
        <div class="section-banner" style="background: linear-gradient(135deg, {theme['start']}, {theme['end']}); color: {theme['text']};">
            <h3 style="color: {theme['text']};">{title}</h3>
            <p>{copy}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def prepare_table_view(df: pd.DataFrame, view_name: str) -> pd.DataFrame:
    config = TABLE_CONFIG.get(view_name, [])
    ordered_columns = [item["field"] for item in config if item["field"] in df.columns]
    remaining_columns = [
        column for column in df.columns if column not in ordered_columns and not str(column).startswith("Unnamed:")
    ]
    return df[ordered_columns + remaining_columns].copy()


def get_column_title_map(view_name: str, columns: list[str]) -> tuple[dict[str, str], str]:
    defaults = {item["field"]: item["header"] for item in TABLE_CONFIG.get(view_name, [])}
    state_key = f"{view_name.lower()}_column_titles"
    existing = st.session_state.get(state_key, {})

    column_title_map = {}
    for column in columns:
        column_title_map[column] = existing.get(column, defaults.get(column, column))

    st.session_state[state_key] = column_title_map
    return column_title_map, state_key


def render_wide_table(
    df: pd.DataFrame,
    view_name: str,
    height: int = 520,
    allow_data_entry: bool = False,
    allow_title_edit: bool = False,
) -> pd.DataFrame:
    if df.empty and len(df.columns) == 0:
        st.info("No columns available for this view.")
        return df

    theme = PAGE_THEMES.get(view_name, {"start": "#083d77", "end": "#12b5ea", "text": "#ffffff"})
    ordered_df = prepare_table_view(df, view_name)
    if allow_data_entry:
        ordered_df = ordered_df.copy()
        ordered_df["__row_index__"] = ordered_df.index
        ordered_df["__pandas_index"] = ordered_df.index.astype(str)

    column_title_map, state_key = get_column_title_map(view_name, list(ordered_df.columns))

    if allow_title_edit:
        with st.expander("Edit column titles", expanded=False):
            updated_map = {}
            for index, column in enumerate(ordered_df.columns):
                if column.startswith("__"):
                    continue
                edited_title = st.text_input(
                    f"{column}",
                    value=column_title_map.get(column, column),
                    key=f"{state_key}_{index}",
                )
                updated_map[column] = edited_title.strip() or column
            for hidden_col in [c for c in ordered_df.columns if c.startswith("__")]:
                updated_map[hidden_col] = hidden_col
            st.session_state[state_key] = updated_map

    display_df = ordered_df.rename(columns=st.session_state[state_key]).fillna("")

    if view_name == "Team":
        # Format all Avg columns as strings so AgGrid types them as text,
        # allowing the pinned row to also use formatted strings without "Invalid Number"
        _name_map = st.session_state[state_key]
        _asc_col = _name_map.get("Avg Secured Conversion %", "Avg Secured Conversion %")
        _aac_col = _name_map.get("Avg Approach Conversion %", "Avg Approach Conversion %")
        _apv_col = _name_map.get("Avg Proposal Value", "Avg Proposal Value")
        if _asc_col in display_df.columns:
            s = pd.to_numeric(display_df[_asc_col], errors="coerce")
            display_df[_asc_col] = s.apply(
                lambda v: "" if pd.isna(v) else (f"{v * 100:.1f}%" if v <= 1 else f"{v:.1f}%")
            )
        if _aac_col in display_df.columns:
            s = pd.to_numeric(display_df[_aac_col], errors="coerce")
            display_df[_aac_col] = s.apply(
                lambda v: "" if pd.isna(v) else (f"{v * 100:.1f}%" if v <= 1 else f"{v:.1f}%")
            )
        if _apv_col in display_df.columns:
            s = pd.to_numeric(display_df[_apv_col], errors="coerce")
            display_df[_apv_col] = s.apply(
                lambda v: "" if pd.isna(v) or v == 0 else f"${v:,.0f}"
            )

    for column in display_df.columns:
        if pd.api.types.is_datetime64_any_dtype(display_df[column]):
            display_df[column] = display_df[column].dt.strftime("%Y-%m-%d").fillna("")

    # Build pinned summary rows for selected views
    pinned_top_rows: list[dict] = []
    if view_name == "Team":
        name_map = st.session_state[state_key]
        pinned_row: dict = {}
        for orig_col in ordered_df.columns:
            display_col = name_map.get(orig_col, orig_col)
            if orig_col == "Team Member":
                pinned_row[display_col] = "TOTAL"
            elif orig_col.startswith("__"):
                pinned_row[display_col] = "__pinned_top__" if orig_col == "__pandas_index" else ""
            elif orig_col == "Avg Secured Conversion %":
                vals = pd.to_numeric(ordered_df[orig_col], errors="coerce").dropna()
                overall = vals.mean() if len(vals) > 0 else 0.0
                pinned_row[display_col] = (f"{overall * 100:.1f}%" if overall > 0 else "-")
            elif orig_col == "Avg Approach Conversion %":
                total_approaches = pd.to_numeric(ordered_df.get("Approaches Made", pd.Series(dtype=float)), errors="coerce").sum()
                total_prospects = pd.to_numeric(ordered_df.get("Prospects", pd.Series(dtype=float)), errors="coerce").sum()
                rate = total_approaches / total_prospects if total_prospects > 0 else 0.0
                pinned_row[display_col] = f"{rate * 100:.1f}%"
            elif orig_col == "Avg Proposal Value":
                total_proposal_val = pd.to_numeric(ordered_df.get("Total Proposal Value", pd.Series(dtype=float)), errors="coerce").sum()
                total_proposals = pd.to_numeric(ordered_df.get("Proposal Requested", pd.Series(dtype=float)), errors="coerce").sum()
                avg_pv = total_proposal_val / total_proposals if total_proposals > 0 else 0.0
                pinned_row[display_col] = f"${avg_pv:,.0f}"
            elif orig_col.startswith("Avg"):
                pinned_row[display_col] = "-"
            else:
                try:
                    total = pd.to_numeric(ordered_df[orig_col], errors="coerce").sum()
                    pinned_row[display_col] = int(total) if float(total).is_integer() else round(float(total), 2)
                except Exception:
                    pinned_row[display_col] = "-"
        pinned_top_rows = [pinned_row]
    elif view_name == "Pipeline":
        name_map = st.session_state[state_key]
        pinned_row: dict = {}
        first_visible_col = next((c for c in ordered_df.columns if not c.startswith("__")), None)

        yes_flag_columns = {
            "Secure Meeting",
            "Quiz Completed",
            "Proposal Sent",
            "Job Secured",
        }
        sum_value_columns = {
            "Existing Fee/$ Value",
            "Proposal Value",
            "Job Secured Value",
            "Additional Work Secured",
        }

        for orig_col in ordered_df.columns:
            display_col = name_map.get(orig_col, orig_col)
            if orig_col.startswith("__"):
                pinned_row[display_col] = "__pinned_top__" if orig_col == "__pandas_index" else ""
                continue

            if orig_col == first_visible_col:
                pinned_row[display_col] = "CALC"
                continue

            series = ordered_df[orig_col]
            if orig_col in yes_flag_columns:
                yes_count = int(series.fillna("").astype(str).str.strip().str.lower().eq("yes").sum())
                pinned_row[display_col] = yes_count
            elif orig_col in sum_value_columns:
                total = float(pd.to_numeric(series, errors="coerce").sum())
                pinned_row[display_col] = int(total) if total.is_integer() else round(total, 2)
            else:
                as_text = series.fillna("").astype(str).str.strip()
                non_blank_count = int(as_text.ne("").sum())
                pinned_row[display_col] = non_blank_count

        pinned_top_rows = [pinned_row]

    remove_header_menu = True
    adjustable_views = {"Pipeline", "Team", "COI"}
    columns_state_key = f"{view_name.lower()}_columns_state"
    persisted_columns_state = st.session_state.get(columns_state_key) if view_name in adjustable_views else None

    builder = GridOptionsBuilder.from_dataframe(display_df)
    min_width = 90 if view_name == "Pipeline" else 80
    builder.configure_default_column(
        resizable=True,
        sortable=True,
        filter=False,
        editable=allow_data_entry,
        minWidth=min_width,
    )
    builder.configure_grid_options(
        animateRows=True,
        suppressDragLeaveHidesColumns=True,
        singleClickEdit=allow_data_entry,
    )
    grid_options = builder.build()

    if pinned_top_rows:
        grid_options["pinnedTopRowData"] = pinned_top_rows

    if allow_data_entry:
        grid_options["getRowId"] = GRID_ROW_ID_JS
    if view_name in adjustable_views:
        grid_options["onGridReady"] = GRID_ON_READY_JS

    if allow_data_entry and view_name in DROPDOWN_OPTIONS:
        name_map = st.session_state[state_key]
        active_opts = st.session_state.get("dropdown_options", DROPDOWN_OPTIONS)
        dropdown_map = active_opts.get(view_name, DROPDOWN_OPTIONS.get(view_name, {}))
        for original_col, values in dropdown_map.items():
            display_col = name_map.get(original_col, original_col)
            for column_def in grid_options.get("columnDefs", []):
                if column_def.get("field") == display_col:
                    column_def["cellEditor"] = "agSelectCellEditor"
                    column_def["cellEditorParams"] = {"values": values}
                    break

    if allow_data_entry:
        row_index_display_name = st.session_state[state_key].get("__row_index__", "__row_index__")
        pandas_index_display_name = st.session_state[state_key].get("__pandas_index", "__pandas_index")
        excel_row_display_name = st.session_state[state_key].get("__excel_row__", "__excel_row__")
        for column_def in grid_options.get("columnDefs", []):
            if column_def.get("field") == row_index_display_name:
                column_def["hide"] = True
                column_def["editable"] = False
            if column_def.get("field") == pandas_index_display_name:
                column_def["hide"] = True
                column_def["editable"] = False
            if column_def.get("field") == excel_row_display_name:
                column_def["hide"] = True
                column_def["editable"] = False

    if remove_header_menu:
        for column_def in grid_options.get("columnDefs", []):
            column_def["sortable"] = True
            column_def["filter"] = False
            column_def["suppressHeaderMenuButton"] = True
            column_def["suppressHeaderContextMenu"] = True
            column_def["wrapHeaderText"] = True
            column_def["autoHeaderHeight"] = True

    if view_name in {"Team", "COI"}:
        for column_def in grid_options.get("columnDefs", []):
            # Keep body cells uniform for reliable manual resize interaction.
            column_def["wrapText"] = False
            column_def["autoHeight"] = False
            column_def["resizable"] = True
            column_def["minWidth"] = 80

    # Keep long workbook-style labels readable without forcing manual resize persistence.
    grid_options["headerHeight"] = 58

    if view_name == "COI":
        # COI columns can contain blanks/text in otherwise numeric-looking fields.
        # Disable type inference to avoid AG Grid showing "Invalid number".
        for column_def in grid_options.get("columnDefs", []):
            column_def["cellDataType"] = False

    cell_white_space = "nowrap"

    custom_css = {
        ".ag-root-wrapper": {
            "border": "2px solid rgba(18, 181, 234, 0.22) !important",
            "border-radius": "18px !important",
            "box-shadow": "0 12px 28px rgba(16, 50, 74, 0.08) !important",
            "overflow": "hidden !important",
        },
        ".ag-header": {
            "background": f"linear-gradient(135deg, {theme['start']}, {theme['end']}) !important",
            "border-bottom": "0 !important",
        },
        ".ag-header-cell": {
            "background": "transparent !important",
            "border-right": "1px solid rgba(255, 255, 255, 0.16) !important",
        },
        ".ag-header-cell-resize": {
            "display": "block !important",
            "opacity": "1 !important",
            "width": "10px !important",
            "right": "-4px !important",
            "cursor": "col-resize !important",
            "z-index": "12 !important",
        },
        ".ag-header-cell-resize::after": {
            "background-color": "rgba(255, 255, 255, 0.62) !important",
            "width": "2px !important",
        },
        ".ag-header-cell-text": {
            "color": f"{theme['text']} !important",
            "font-weight": "700 !important",
        },
        ".ag-cell": {
            "white-space": f"{cell_white_space} !important",
        },
        ".ag-cell-value": {
            "white-space": f"{cell_white_space} !important",
            "overflow-wrap": "anywhere !important",
            "word-break": "break-word !important",
            "line-height": "1.25 !important",
        },
        ".ag-cell-wrapper": {
            "white-space": f"{cell_white_space} !important",
            "overflow-wrap": "anywhere !important",
            "word-break": "break-word !important",
        },
        ".ag-body-horizontal-scroll": {
            "height": "14px !important",
        },
        ".ag-floating-top": {
            "background": "linear-gradient(135deg, #fff3cd, #ffe69c) !important",
            "border-bottom": "2px solid rgba(255, 193, 7, 0.6) !important",
        },
        ".ag-floating-top .ag-cell": {
            "font-weight": "700 !important",
            "color": "#4a2a00 !important",
        },
    }

    if allow_data_entry:
        st.caption("Data entry is enabled on this page. Click a cell to edit it, drag headers to reorder columns, and scroll sideways as needed.")
    else:
        st.caption("Summary view only. Drag headers left or right to reorder columns, and scroll sideways across the grid.")

    grid_update_on = None
    if view_name == "Pipeline" and allow_data_entry:
        # Include cellValueChanged for editable pages, plus calmer layout events for width/order persistence.
        grid_update_on = ["cellValueChanged", ("columnResized", 450), ("columnMoved", 450)]
    elif view_name == "COI" and allow_data_entry:
        grid_update_on = ["cellValueChanged", ("columnResized", 450), ("columnMoved", 450)]
    elif view_name == "Team":
        grid_update_on = [("columnResized", 450), ("columnMoved", 450)]

    grid_response = AgGrid(
        display_df,
        gridOptions=grid_options,
        custom_css=custom_css,
        enable_enterprise_modules=False,
        allow_unsafe_jscode=(allow_data_entry or view_name in adjustable_views),
        columns_state=persisted_columns_state,
        fit_columns_on_grid_load=False,
        height=height,
        theme="streamlit",
        should_grid_return=GRID_SHOULD_RETURN_JS if view_name in adjustable_views else None,
        update_mode=GridUpdateMode.NO_UPDATE if grid_update_on else (GridUpdateMode.VALUE_CHANGED if allow_data_entry else GridUpdateMode.NO_UPDATE),
        update_on=grid_update_on or [],
        key=f"{view_name.lower()}_grid_v7",
    )

    if view_name in adjustable_views and grid_response:
        latest_columns_state = getattr(grid_response, "columns_state", None)
        if isinstance(latest_columns_state, list) and latest_columns_state:
            latest_signature = json.dumps(latest_columns_state, sort_keys=True, default=str)
            previous_signature = st.session_state.get(f"{columns_state_key}_sig")

            if latest_signature != previous_signature:
                st.session_state[columns_state_key] = latest_columns_state
                st.session_state[f"{columns_state_key}_sig"] = latest_signature
                set_persistent_columns_state(view_name, latest_columns_state)

                event_data = getattr(grid_response, "event_data", None) or {}
                trigger_name = event_data.get("streamlitRerunEventTriggerName")
                if trigger_name in {"columnResized", "columnMoved"}:
                    toast_key = f"{columns_state_key}_last_toast_time"
                    now_ts = time.time()
                    last_toast_ts = float(st.session_state.get(toast_key, 0.0) or 0.0)
                    # Debounce to avoid noisy toasts while the user is dragging column widths.
                    if now_ts - last_toast_ts >= 1.2:
                        st.toast(f"{view_name} column layout saved", icon="âœ…")
                        st.session_state[toast_key] = now_ts

    if allow_data_entry and grid_response and "data" in grid_response:
        updated_df = pd.DataFrame(grid_response["data"])
        reverse_map = {}
        for original_name in ordered_df.columns:
            display_name = st.session_state[state_key].get(original_name, original_name)
            reverse_map.setdefault(display_name, original_name)
        updated_df = updated_df.rename(columns={col: reverse_map.get(col, col) for col in updated_df.columns})
        if "__row_index__" in updated_df.columns:
            updated_df["__row_index__"] = pd.to_numeric(updated_df["__row_index__"], errors="coerce")
            updated_df = updated_df.dropna(subset=["__row_index__"])
            updated_df.index = updated_df["__row_index__"].astype(int)
            updated_df = updated_df.drop(columns=["__row_index__"])
        if "__pandas_index" in updated_df.columns:
            updated_df = updated_df.drop(columns=["__pandas_index"])
        return updated_df

    return display_df


def to_excel_cell_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def append_blank_rows(df: pd.DataFrame, row_count: int) -> pd.DataFrame:
    if row_count <= 0:
        return df
    blank_rows = pd.DataFrame(
        "",
        index=range(NEW_ROW_SENTINEL, NEW_ROW_SENTINEL + row_count),
        columns=df.columns,
    )
    return pd.concat([df, blank_rows], axis=0)


def apply_new_row_defaults(df: pd.DataFrame, view_name: str) -> pd.DataFrame:
    if df.empty:
        return df
    defaults = DEFAULT_NEW_ROW_VALUES.get(view_name, {})
    if not defaults:
        return df

    result = df.copy()
    new_row_mask = result.index.to_series().apply(lambda x: int(x) >= NEW_ROW_SENTINEL)
    for column, default_value in defaults.items():
        if column not in result.columns:
            continue
        col_text = result[column].astype(str).str.strip()
        blank_mask = col_text.eq("") | col_text.eq("nan")
        set_mask = new_row_mask & blank_mask
        if set_mask.any():
            result.loc[set_mask, column] = default_value
    return result


def save_dropdown_options_to_lists(workbook_path: str, options: dict[str, dict[str, list[str]]]) -> tuple[int, list[str]]:
    wb = load_workbook(workbook_path, data_only=False)
    if "Lists" not in wb.sheetnames:
        wb.close()
        raise ValueError("Workbook does not contain a 'Lists' sheet.")

    ws = wb["Lists"]
    updated_cells = 0
    truncated_fields: list[str] = []

    for view_name, field_map in LISTS_RANGE_MAP.items():
        source_fields = options.get(view_name, {})
        for field_name, (start_row, end_row) in field_map.items():
            values = [str(v).strip() for v in source_fields.get(field_name, []) if str(v).strip()]
            capacity = end_row - start_row + 1
            if len(values) > capacity:
                values = values[:capacity]
                truncated_fields.append(f"{view_name} - {field_name}")

            for row in range(start_row, end_row + 1):
                ws.cell(row=row, column=2, value=None)
                updated_cells += 1

            for idx, value in enumerate(values):
                ws.cell(row=start_row + idx, column=2, value=value)
                updated_cells += 1

    wb.save(workbook_path)
    wb.close()
    return updated_cells, truncated_fields


def write_sheet_updates(
    workbook,
    workbook_path: str,
    sheet_name: str,
    edited_df: pd.DataFrame,
    editable_fields: list[str],
    header_markers: list[str],
) -> int:
    if edited_df.empty:
        return 0

    header_row_index = find_header_row(workbook_path, sheet_name, header_markers)
    header_excel_row = header_row_index + 1

    worksheet = workbook[sheet_name]
    header_map = {}
    for col_idx in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(row=header_excel_row, column=col_idx).value
        if header_value is not None:
            header_map[str(header_value).strip()] = col_idx

    updated_cells = 0
    valid_fields = [field for field in editable_fields if field in edited_df.columns and field in header_map]

    # Determine append start for newly added rows.
    next_append_excel_row = worksheet.max_row + 1
    for row_index, row_values in edited_df.iterrows():
        row_idx = int(row_index)
        if row_idx >= NEW_ROW_SENTINEL:
            excel_row = next_append_excel_row
            next_append_excel_row += 1
        else:
            preserved_excel_row = pd.to_numeric(pd.Series([row_values.get("__excel_row__")]), errors="coerce").iloc[0]
            if pd.notna(preserved_excel_row):
                excel_row = int(preserved_excel_row)
            else:
                excel_row = header_excel_row + 1 + row_idx
        for field in valid_fields:
            excel_col = header_map[field]
            worksheet.cell(row=excel_row, column=excel_col).value = to_excel_cell_value(row_values[field])
            updated_cells += 1

    return updated_cells


def save_activity_changes(
    workbook_path: str,
    pipeline_updates: pd.DataFrame | None = None,
    coi_updates: pd.DataFrame | None = None,
) -> int:
    create_workbook_backup(workbook_path)
    try:
        workbook = load_workbook(workbook_path)
    except PermissionError as exc:
        raise PermissionError("Could not open the workbook for saving. Close it in Excel and try again.") from exc
    total_updated_cells = 0

    if pipeline_updates is not None and not pipeline_updates.empty:
        pipeline_fields = [column for column in pipeline_updates.columns if not str(column).startswith("__")]
        total_updated_cells += write_sheet_updates(
            workbook,
            workbook_path,
            "Sales Activity",
            pipeline_updates,
            pipeline_fields,
            HEADER_MARKERS["Sales Activity"],
        )

    if coi_updates is not None and not coi_updates.empty:
        coi_fields = [item["field"] for item in TABLE_CONFIG["COI"]]
        total_updated_cells += write_sheet_updates(
            workbook,
            workbook_path,
            "COI Development",
            coi_updates,
            coi_fields,
            HEADER_MARKERS["COI Development"],
        )

    try:
        workbook.save(workbook_path)
    except PermissionError as exc:
        raise PermissionError("Could not save the workbook. Close it in Excel and try again.") from exc
    finally:
        workbook.close()
    return total_updated_cells


def build_sidebar(sales_df: pd.DataFrame) -> tuple[pd.DataFrame, str, bool]:
    # --- Access level ---
    st.sidebar.header("Access level")
    if "is_manager" not in st.session_state:
        st.session_state.is_manager = False

    role_label = "Firm Manager" if st.session_state.is_manager else "Advisor"
    st.sidebar.markdown(f"Signed in as: **{role_label}**")

    if st.session_state.is_manager:
        if st.sidebar.button("Sign out to Advisor", use_container_width=True):
            st.session_state.is_manager = False
            st.session_state["current_page"] = "Home"
            st.rerun()

        with st.sidebar.expander("Update manager password", expanded=False):
            new_pwd = st.text_input("New password", type="password", key="manager_pwd_new")
            confirm_pwd = st.text_input("Confirm new password", type="password", key="manager_pwd_confirm")
            if st.button("Save manager password", key="manager_pwd_save", use_container_width=True):
                if not new_pwd.strip():
                    st.error("Password cannot be blank.")
                elif new_pwd != confirm_pwd:
                    st.error("Passwords do not match.")
                else:
                    config_data = load_app_config()
                    config_data["manager_password"] = new_pwd
                    save_app_config(config_data)
                    st.success("Manager password saved.")
    else:
        with st.sidebar.expander("Sign in as Firm Manager", expanded=False):
            manager_password = get_manager_password()
            pwd = st.text_input("Password", type="password", key="manager_pwd_input")
            if st.button("Sign in", key="manager_sign_in", use_container_width=True):
                if not manager_password:
                    st.error("Manager sign-in is not configured. Save a manager password below or set SALES_TRACKER_MANAGER_PASSWORD.")
                elif pwd == manager_password:
                    st.session_state.is_manager = True
                    st.rerun()
                else:
                    st.error("Incorrect password.")

            if not manager_password:
                st.caption("No manager password is configured yet.")
                initial_pwd = st.text_input("Set manager password", type="password", key="manager_pwd_initial")
                confirm_initial_pwd = st.text_input("Confirm manager password", type="password", key="manager_pwd_initial_confirm")
                if st.button("Create manager password", key="manager_pwd_create", use_container_width=True):
                    if not initial_pwd.strip():
                        st.error("Password cannot be blank.")
                    elif initial_pwd != confirm_initial_pwd:
                        st.error("Passwords do not match.")
                    else:
                        config_data = load_app_config()
                        config_data["manager_password"] = initial_pwd
                        save_app_config(config_data)
                        st.success("Manager password saved. You can sign in now.")

    st.sidebar.divider()

    st.sidebar.header("Data source")
    workbook_input = st.sidebar.text_input("Workbook path", value=str(DEFAULT_WORKBOOK))
    workbook_ok, workbook_error = validate_workbook_path(workbook_input)
    if not workbook_ok and workbook_error:
        st.sidebar.error(workbook_error)

    st.sidebar.header("Filters")

    staff_options = sorted(
        [x for x in sales_df.get("Lead Staff (Client Manager)", pd.Series(dtype=str)).dropna().unique().tolist() if str(x).strip()]
    )
    selected_staff = st.sidebar.multiselect("Lead staff", options=staff_options, default=staff_options)

    status_options = sorted(
        [x for x in sales_df.get("Prospect Status", pd.Series(dtype=str)).dropna().unique().tolist() if str(x).strip()]
    )
    selected_status = st.sidebar.multiselect("Prospect status", options=status_options, default=status_options)

    date_col = "Approach Date"
    if date_col in sales_df.columns and sales_df[date_col].notna().any():
        min_date = sales_df[date_col].min().date()
        max_date = sales_df[date_col].max().date()
        selected_dates = st.sidebar.date_input("Approach date range", value=(min_date, max_date), min_value=min_date, max_value=max_date)
    else:
        selected_dates = None

    filtered = sales_df.copy()
    if selected_staff and "Lead Staff (Client Manager)" in filtered.columns:
        filtered = filtered[filtered["Lead Staff (Client Manager)"].isin(selected_staff)]
    if selected_status and "Prospect Status" in filtered.columns:
        filtered = filtered[filtered["Prospect Status"].isin(selected_status)]

    if selected_dates and len(selected_dates) == 2 and "Approach Date" in filtered.columns:
        start, end = selected_dates
        filtered = filtered[(filtered["Approach Date"].dt.date >= start) & (filtered["Approach Date"].dt.date <= end)]

    return filtered, workbook_input, st.session_state.is_manager


def kpi_row(filtered_sales: pd.DataFrame) -> None:
    approaches = int(filtered_sales.get("Approach Style", pd.Series(dtype=str)).fillna("").astype(str).str.strip().ne("").sum())
    meetings = int((filtered_sales.get("Secure Meeting", pd.Series(dtype=str)).fillna("").astype(str).str.lower() == "yes").sum())
    proposals = int((filtered_sales.get("Proposal Sent", pd.Series(dtype=str)).fillna("").astype(str).str.lower() == "yes").sum())
    secured_value = float(filtered_sales.get("Job Secured Value", pd.Series(dtype=float)).sum())

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Approaches", f"{approaches:,}")
    c2.metric("Meetings secured", f"{meetings:,}")
    c3.metric("Proposals sent", f"{proposals:,}")
    c4.metric("Work secured", f"${secured_value:,.0f}")


def charts(filtered_sales: pd.DataFrame, coi_df: pd.DataFrame) -> None:
    left, right = st.columns(2)

    if "Lead Staff (Client Manager)" in filtered_sales.columns:
        by_staff = (
            filtered_sales.groupby("Lead Staff (Client Manager)", dropna=False)["Job Secured Value"]
            .sum()
            .reset_index()
            .sort_values("Job Secured Value", ascending=False)
        )
        fig_staff = px.bar(
            by_staff,
            x="Lead Staff (Client Manager)",
            y="Job Secured Value",
            title="Work secured by staff",
            color="Job Secured Value",
            color_continuous_scale="Tealgrn",
        )
        left.plotly_chart(fig_staff, use_container_width=True)

    if "Approach Date" in filtered_sales.columns:
        monthly = filtered_sales.dropna(subset=["Approach Date"]).copy()
        if not monthly.empty:
            monthly["Month"] = monthly["Approach Date"].dt.to_period("M").astype(str)
            monthly_chart = monthly.groupby("Month", as_index=False)["Job Secured Value"].sum()
            fig_month = px.line(
                monthly_chart,
                x="Month",
                y="Job Secured Value",
                title="Monthly secured value trend",
                markers=True,
            )
            right.plotly_chart(fig_month, use_container_width=True)

    row2_left, row2_right = st.columns(2)

    if "Prospect Status" in filtered_sales.columns:
        status_counts = filtered_sales["Prospect Status"].fillna("Unknown").value_counts().reset_index()
        status_counts.columns = ["Prospect Status", "Count"]
        fig_status = px.pie(
            status_counts,
            names="Prospect Status",
            values="Count",
            title="Prospect status mix",
            hole=0.45,
        )
        row2_left.plotly_chart(fig_status, use_container_width=True)

    if "Prospect Source" in filtered_sales.columns:
        source_counts = (
            filtered_sales["Prospect Source"]
            .fillna("").astype(str).str.strip()
            .replace("", "Unknown")
            .value_counts()
            .reset_index()
        )
        source_counts.columns = ["Prospect Source", "Count"]
        source_counts = source_counts[source_counts["Prospect Source"] != "Unknown"]
        if not source_counts.empty:
            fig_source = px.bar(
                source_counts,
                x="Prospect Source",
                y="Count",
                title="New prospect source",
                color="Count",
                color_continuous_scale="Teal",
            )
            fig_source.update_layout(xaxis_tickangle=-30)
            row2_right.plotly_chart(fig_source, use_container_width=True)

    if "Industry" in coi_df.columns:
        coi_industry = coi_df.copy()
        coi_industry["Industry"] = coi_industry["Industry"].astype(str).str.strip()
        coi_industry = coi_industry[coi_industry["Industry"].ne("") & coi_industry["Industry"].ne("nan")]
        if not coi_industry.empty:
            industry_counts = (
                coi_industry["Industry"]
                .value_counts()
                .rename_axis("Industry")
                .reset_index(name="Relationships")
            )
            fig_coi_industry = px.bar(
                industry_counts,
                x="Industry",
                y="Relationships",
                title="COI relationships by industry",
                color="Relationships",
                color_continuous_scale="Magenta",
            )
            st.plotly_chart(fig_coi_industry, use_container_width=True)


def build_blog_draft(
    topic: str,
    audience: str,
    objective: str,
    tone: str,
    length: str,
    principles: list[dict[str, object]],
    cta: str,
    word_range: tuple[int, int] | None = None,
) -> str:
    sections_by_length = {
        "Short": 2,
        "Medium": 3,
        "Long": 3,
    }
    section_count = sections_by_length.get(length, 3)

    normalized: list[dict[str, object]] = []
    for p in principles:
        title_text = str(p.get("title", "")).strip()
        details_raw = p.get("details", [])
        details = [str(item).strip() for item in details_raw if str(item).strip()][:3] if isinstance(details_raw, list) else []
        if title_text or details:
            normalized.append({
                "title": title_text or "Practical focus area",
                "details": details,
            })

    if not normalized:
        normalized = [
            {
                "title": "Market context and immediate signals",
                "details": [
                    "What is changing now and why it matters",
                    "Where uncertainty is highest",
                    "Which indicators deserve close attention",
                ],
            },
            {
                "title": "Actions clients can take this month",
                "details": [
                    "A practical first step",
                    "How to sequence decisions",
                    "How to avoid overreaction",
                ],
            },
            {
                "title": "How to stay accountable to a plan",
                "details": [
                    "What to review regularly",
                    "When to adjust your strategy",
                    "How to keep decisions objective",
                ],
            },
        ]

    selected_principles = normalized[:section_count]

    tone_voice = {
        "Professional": "clear and practical",
        "Friendly": "conversational and supportive",
        "Confident": "direct and decisive",
        "Educational": "explanatory and step-by-step",
    }.get(tone, "clear and practical")

    title = f"{topic}: A Practical Guide for {audience}"
    intro = (
        f"# {title}\n\n"
        f"In this {tone_voice} post, we will break down {topic.lower()} and connect it to one core goal: {objective.lower()}. "
        f"If you are part of {audience.lower()}, the aim is to leave you with specific, useful actions you can take right away."
    )

    section_blocks: list[str] = []
    for idx, principle in enumerate(selected_principles, start=1):
        p_title = str(principle.get("title", f"Principle {idx}")).strip() or f"Principle {idx}"
        p_details = principle.get("details", [])
        detail_lines: list[str] = []
        if isinstance(p_details, list) and p_details:
            for detail in p_details[:3]:
                d = str(detail).strip()
                if not d:
                    continue
                detail_lines.append(
                    f"- **{d}**: This is where we translate the idea into a practical decision point and a realistic next step."
                )
        if not detail_lines:
            detail_lines.append(
                "- **Key takeaway**: Focus on one clear action and measure the outcome before making your next move."
            )

        section_blocks.append(
            "\n\n".join(
                [
                    f"## Principle {idx}: {p_title}",
                    f"This principle supports {objective.lower()} by focusing on decisions that are clear, timely, and measurable.",
                    "\n".join(detail_lines),
                    "When applied consistently, this principle helps reduce noise and improves confidence in your planning process.",
                ]
            )
        )

    closing = (
        "## Final Takeaway\n"
        f"Strong outcomes usually come from a few well-executed principles, not from reacting to every headline. "
        f"If you would like tailored guidance for your own position, {cta}."
    )

    draft_text = "\n\n".join([intro, *section_blocks, closing])
    if word_range:
        draft_text = fit_text_to_word_range(draft_text, word_range[0], word_range[1])
    return draft_text


def build_final_blog_post_from_outline(
    outline_text: str,
    topic: str,
    audience: str,
    objective: str,
    tone: str,
    cta: str,
    word_range: tuple[int, int] | None = None,
    polish_level: str = "Strong",
    rewrite_intensity: str = "Editorial",
    style_examples: list[str] | None = None,
    style_strength: str = "Balanced",
    style_example_names: list[str] | None = None,
) -> str:
    def detail_to_sentence(detail: str, tone_value: str, detail_index: int) -> str:
        d = detail.strip().lower().rstrip(".?!")
        if not d:
            return ""

        if "risk" in d:
            return f"Treat {d} as something to actively monitor so it does not quietly compound over time."
        if "opportun" in d:
            return f"Look for practical ways to capture {d} without stretching beyond your capacity to execute well."
        if "action" in d or "next step" in d:
            return f"Turn {d} into a dated commitment so momentum is created immediately."
        if "signal" in d or "indicator" in d:
            return f"Use {d} as a recurring checkpoint to separate short-term noise from meaningful change."
        if "mistake" in d:
            return f"A simple way to avoid {d} is to define decision criteria before pressure builds."
        if "follow-up" in d or "review" in d:
            return f"Schedule {d} now, because consistency matters more than intensity in a single week."

        fallback_templates = {
            "Confident": [
                "Prioritize {d} and execute with discipline; progress follows decisive, repeated action.",
                "Treat {d} as a non-negotiable focus area and assign ownership immediately.",
                "Use {d} as a practical decision trigger rather than waiting for perfect conditions.",
            ],
            "Friendly": [
                "A useful place to start is {d}, then build confidence with one clear follow-through step.",
                "Start with {d} and keep it simple: one decision, one action, one review point.",
                "Make {d} easier to act on by linking it to your next scheduled check-in.",
            ],
            "Educational": [
                "Think of {d} as a practical lever: when used consistently, it improves decision quality over time.",
                "Use {d} as a repeatable input in your decision process so outcomes become more predictable.",
                "Framing {d} as a measurable criterion can improve clarity and reduce bias in choices.",
            ],
            "Professional": [
                "Translate {d} into one clear action for this period, then review the result objectively.",
                "Use {d} to sharpen priorities so decisions stay aligned to outcomes that matter.",
                "Build {d} into your decision cadence to improve consistency and reduce drift.",
            ],
        }
        tone_key = tone_value if tone_value in fallback_templates else "Professional"
        template = fallback_templates[tone_key][detail_index % len(fallback_templates[tone_key])]
        return template.format(d=d)

    lines = [line.strip() for line in (outline_text or "").splitlines() if line.strip()]

    title = topic.strip() if topic.strip() else "Client Update"
    for line in lines:
        if re.match(r"^#\s+", line):
            candidate = re.sub(r"^#\s+", "", line).strip()
            if candidate:
                title = candidate
                break

    sections: list[dict[str, object]] = []
    current: dict[str, object] | None = None
    skip_prefixes = (
        "this principle supports",
        "when applied consistently",
        "key takeaway",
        "final takeaway",
    )

    for raw in lines:
        lowered = raw.lower()
        if lowered.startswith("## principle"):
            if current:
                sections.append(current)
            heading = raw.split(":", 1)[1].strip() if ":" in raw else raw.replace("##", "").strip()
            current = {"heading": heading, "details": [], "notes": []}
            continue

        if raw.startswith("-"):
            bullet = raw.lstrip("- ").strip()
            bullet = re.sub(r"\*\*(.*?)\*\*", r"\1", bullet)
            bullet = bullet.split(":", 1)[0].strip() if ":" in bullet else bullet
            if current and bullet:
                details = current.get("details", [])
                if isinstance(details, list) and bullet not in details:
                    details.append(bullet)
                    current["details"] = details
            continue

        if any(lowered.startswith(prefix) for prefix in skip_prefixes):
            continue

        if current:
            notes = current.get("notes", [])
            if isinstance(notes, list) and raw not in notes:
                notes.append(raw)
                current["notes"] = notes

    if current:
        sections.append(current)

    if not sections:
        sections = [
            {
                "heading": "Context and relevance",
                "details": ["what is changing", "why it matters now"],
                "notes": ["Use this section to connect recent changes to practical decisions."],
            },
            {
                "heading": "Practical actions",
                "details": ["where to prioritize", "what to review this month"],
                "notes": ["Focus on decisions that are realistic and measurable."],
            },
        ]

    def human_join(items: list[str]) -> str:
        cleaned = [i.strip() for i in items if i.strip()]
        if not cleaned:
            return ""
        if len(cleaned) == 1:
            return cleaned[0]
        if len(cleaned) == 2:
            return f"{cleaned[0]} and {cleaned[1]}"
        return f"{', '.join(cleaned[:-1])}, and {cleaned[-1]}"

    tone_open_base = {
        "Professional": "This article gives a practical view you can apply immediately.",
        "Friendly": "This article keeps things clear, practical, and easy to use in the real world.",
        "Confident": "This article focuses on decisive actions that create momentum.",
        "Educational": "This article explains the key ideas and shows how to apply them step by step.",
    }.get(tone, "This article gives a practical view you can apply immediately.")

    polish_mode = str(polish_level or "Strong").strip().title()
    if polish_mode not in {"Standard", "Strong", "Premium"}:
        polish_mode = "Strong"

    rewrite_mode = str(rewrite_intensity or "Editorial").strip().title()
    if rewrite_mode not in {"Conservative", "Editorial", "Publish-Ready"}:
        rewrite_mode = "Editorial"

    style_profile = build_style_profile(
        style_examples,
        style_strength=style_strength,
        style_example_names=style_example_names,
    )
    starter_phrases = [str(v).strip() for v in style_profile.get("starter_phrases", []) if str(v).strip()]
    signature_terms = [str(v).strip() for v in style_profile.get("signature_terms", []) if str(v).strip()]
    avg_sentence_words = int(style_profile.get("avg_sentence_words", 18) or 18)

    if polish_mode == "Standard":
        tone_open = tone_open_base
        section_openers = [
            "{transition} {heading} helps keep decisions grounded in what matters most.",
            "{transition} the focus shifts to {heading_lower}, where strategy becomes practical action.",
            "{transition} {heading} supports consistency so progress is maintained over time.",
        ]
    elif polish_mode == "Premium":
        tone_open = (
            f"{tone_open_base} The focus here is not just understanding change, "
            f"but converting insight into confident decisions with measurable progress."
        )
        section_openers = [
            "{transition} {heading} creates the strategic frame for disciplined decision-making.",
            "{transition} attention turns to {heading_lower}, where intent must be translated into execution quality.",
            "{transition} {heading} becomes the operating rhythm that protects momentum and reinforces accountability.",
        ]
    else:
        tone_open = f"{tone_open_base} The emphasis is clarity, relevance, and action."
        section_openers = [
            "{transition} {heading} is the anchor for consistent, practical choices.",
            "{transition} the focus moves to {heading_lower}, where plans are converted into action.",
            "{transition} {heading} strengthens execution so momentum and accountability stay intact.",
        ]

    if starter_phrases:
        transitions = [
            f"{starter_phrases[0]},",
            f"{starter_phrases[1]}," if len(starter_phrases) > 1 else "Next,",
            f"{starter_phrases[2]}," if len(starter_phrases) > 2 else "From there,",
        ]
    else:
        transitions = ["First,", "Next,", "From there,"]

    cadence_line = (
        "Keep each point concise and direct so readers can act quickly."
        if avg_sentence_words <= 16
        else "Use fuller context in each section while keeping the practical takeaway explicit."
    )
    bridge_lines = [
        "Viewed together, these choices create clarity and reduce reactive decision-making.",
        "The value comes from linking each decision back to your broader objective and timeline.",
        "Over time, this approach turns uncertain conditions into a manageable decision process.",
    ]
    if signature_terms:
        bridge_lines[1] = (
            f"Keep attention on {human_join(signature_terms[:3])} so every action stays connected to your core objective."
        )

    max_details_by_mode = {
        "Conservative": 1,
        "Editorial": 2,
        "Publish-Ready": 3,
    }
    detail_limit = max_details_by_mode[rewrite_mode]

    paragraphs: list[str] = []
    paragraphs.append(
        f"{title}\n\n"
        f"{tone_open} {cadence_line} If you are part of {audience.lower()}, the goal is simple: {objective.lower()} through clear, practical decisions."
    )

    for idx, section in enumerate(sections[:3]):
        heading = str(section.get("heading", "Practical focus")).strip() or "Practical focus"
        details = section.get("details", [])
        notes = section.get("notes", [])

        detail_summary = ""
        detail_expansion = ""
        if isinstance(details, list) and details:
            clean_details = [str(d).strip() for d in details if str(d).strip()]
            if len(clean_details) == 1:
                detail_summary = f"A useful starting point is {clean_details[0].lower()}."
            elif len(clean_details) >= 2:
                detail_summary = f"In practical terms, prioritize {human_join([d.lower() for d in clean_details[:detail_limit]])}."

            expansion_lines = [detail_to_sentence(d, tone, i) for i, d in enumerate(clean_details[:detail_limit])]
            expansion_lines = [line for line in expansion_lines if line]
            if expansion_lines:
                detail_expansion = " ".join(expansion_lines)

        note_text = ""
        if isinstance(notes, list) and notes:
            candidates = [str(n).strip() for n in notes if str(n).strip()]
            if candidates:
                note_text = candidates[0]

        transition = transitions[idx % len(transitions)]
        opener = section_openers[idx % len(section_openers)].format(
            transition=transition,
            heading=heading,
            heading_lower=heading.lower(),
        )
        sentence_parts = [opener]
        if detail_summary:
            sentence_parts.append(detail_summary)
        if note_text:
            sentence_parts.append(note_text)
        if detail_expansion:
            sentence_parts.append(detail_expansion)
        sentence_parts.append(bridge_lines[idx % len(bridge_lines)])
        paragraph = " ".join([part.strip() for part in sentence_parts if part.strip()])
        paragraphs.append(paragraph.strip())

    if rewrite_mode == "Publish-Ready":
        paragraphs.append(
            "Across all three principles, the central theme is discipline under uncertainty: "
            "define what matters, commit to a practical cadence, and let evidence guide course-corrections."
        )

    closing_lines = {
        "Standard": "The objective is steady progress, not reactionary change.",
        "Strong": "The objective is steady progress through deliberate, well-timed decisions rather than reactive moves.",
        "Premium": "The objective is sustained progress built on disciplined choices, clear priorities, and consistent execution over time.",
    }
    cta_text = str(cta or "").strip().rstrip(".")
    if not cta_text:
        cta_text = "book a short strategy call"
    section_names = [str(s.get("heading", "")).strip().lower() for s in sections[:3] if str(s.get("heading", "")).strip()]
    if section_names:
        summary_line = f"In short, strong outcomes come from managing {human_join(section_names)} with consistency and intent."
    else:
        summary_line = "In short, strong outcomes come from consistent execution of a few high-value decisions."

    paragraphs.append(
        f"{summary_line} {closing_lines[polish_mode]} If you would like tailored guidance for your own situation, {cta_text}."
    )

    final_text = "\n\n".join(paragraphs)
    if word_range:
        final_text = fit_text_to_word_range(final_text, word_range[0], word_range[1])
    return final_text


def count_words(text: str) -> int:
    return len(re.findall(r"\b[\w'-]+\b", text or ""))


def fit_text_to_word_range(text: str, min_words: int, max_words: int) -> str:
    min_words = max(10, int(min_words))
    max_words = max(min_words, int(max_words))
    words = (text or "").split()

    def trim_to_complete_sentences(input_text: str, limit_words: int) -> str:
        sentences = [s.strip() for s in re.split(r"(?<=[.!?])\s+", input_text.strip()) if s.strip()]
        if not sentences:
            raw = " ".join(input_text.split()[:limit_words]).strip()
            return raw + ("." if raw and raw[-1] not in ".!?" else "")

        kept: list[str] = []
        running = 0
        for sentence in sentences:
            s_words = len(sentence.split())
            if kept and running + s_words > limit_words:
                break
            if not kept and s_words > limit_words:
                clipped = " ".join(sentence.split()[:limit_words]).strip()
                return clipped + ("." if clipped and clipped[-1] not in ".!?" else "")
            kept.append(sentence)
            running += s_words

        output = " ".join(kept).strip()
        if not output:
            output = " ".join(input_text.split()[:limit_words]).strip()
        if output and output[-1] not in ".!?":
            output += "."
        return output

    if len(words) > max_words:
        return trim_to_complete_sentences(text, max_words)

    if len(words) < min_words:
        filler_blocks = [
            "\n\nTo make this practical, choose one action from the principles above and schedule a clear review date.",
            "\n\nAs conditions change, use these principles as a filter so every decision stays aligned to your long-term objective.",
            "\n\nMost clients gain momentum when they focus on consistency first, then optimize details once the core plan is working.",
        ]
        out = text
        idx = 0
        while len(out.split()) < min_words:
            out += filler_blocks[idx % len(filler_blocks)]
            idx += 1
        if len(out.split()) > max_words:
            out = trim_to_complete_sentences(out, max_words)
        return out

    return text


def build_outreach_messages(blog_text: str, cta: str) -> dict[str, str]:
    lines = [line.strip() for line in blog_text.splitlines() if line.strip()]
    title_line = next((line.lstrip("# ").strip() for line in lines if line.startswith("#")), "New update from your advisory team")

    summary_chunks = [line for line in lines if not line.startswith("#") and not line.startswith("###")]
    summary = " ".join(summary_chunks)[:380].strip()
    if not summary:
        summary = "We have prepared a practical update that can help you plan your next steps with more confidence."

    email_subject = f"{title_line}"
    email_body = (
        f"Hi,\n\n"
        f"{summary}\n\n"
        f"{cta}.\n\n"
        f"Best regards,\nYour advisory team"
    )
    sms_text = f"{title_line} - {summary[:180]} {cta}."
    linkedin_post = f"{title_line}\n\n{summary}\n\n{cta}. #Advisory #ClientUpdate"
    facebook_post = f"{title_line}\n\n{summary}\n\n{cta}."
    instagram_post = f"{title_line}\n\n{summary[:220]}\n\n{cta}. #finance #planning #advice"

    return {
        "email_subject": email_subject,
        "email_body": email_body,
        "sms": sms_text,
        "linkedin": linkedin_post,
        "facebook": facebook_post,
        "instagram": instagram_post,
    }


def build_campaign_csv(recipients: pd.DataFrame, channel: str, messages: dict[str, str]) -> pd.DataFrame:
    if channel == "email":
        out = recipients[["Email"]].copy() if "Email" in recipients.columns else pd.DataFrame(columns=["Email"])
        out = out.rename(columns={"Email": "To"})
        out["Subject"] = messages["email_subject"]
        out["Body"] = messages["email_body"]
        return out

    if channel == "sms":
        out = recipients[["Contact Phone #"]].copy() if "Contact Phone #" in recipients.columns else pd.DataFrame(columns=["Contact Phone #"])
        out = out.rename(columns={"Contact Phone #": "Phone"})
        out["Message"] = messages["sms"]
        return out

    # Social channels are typically one public post, not per-recipient direct messages.
    return pd.DataFrame(
        {
            "Channel": [channel.title()],
            "Post Text": [messages[channel]],
        }
    )


def get_blog_draft_history() -> list[dict]:
    config_data = load_app_config()
    history = config_data.get("blog_draft_history", [])
    if not isinstance(history, list):
        return []
    return [item for item in history if isinstance(item, dict)]


def save_blog_draft_to_history(entry: dict, max_items: int = 40) -> int:
    config_data = load_app_config()
    history = config_data.get("blog_draft_history", [])
    if not isinstance(history, list):
        history = []
    history = [item for item in history if isinstance(item, dict)]
    history.insert(0, entry)
    config_data["blog_draft_history"] = history[:max_items]
    save_app_config(config_data)
    return len(config_data["blog_draft_history"])


def delete_blog_draft_from_history(index: int) -> bool:
    config_data = load_app_config()
    history = config_data.get("blog_draft_history", [])
    if not isinstance(history, list):
        return False
    history = [item for item in history if isinstance(item, dict)]
    if index < 0 or index >= len(history):
        return False
    history.pop(index)
    config_data["blog_draft_history"] = history
    save_app_config(config_data)
    return True


def get_blog_style_examples() -> list[dict]:
    config_data = load_app_config()
    examples = config_data.get("blog_style_examples", [])
    if not isinstance(examples, list):
        return []
    cleaned: list[dict] = []
    for item in examples:
        if not isinstance(item, dict):
            continue
        title = str(item.get("title", "")).strip()
        content = str(item.get("content", "")).strip()
        source_name = str(item.get("source_name", "")).strip()
        if not content:
            continue
        cleaned.append(
            {
                "title": title or "Untitled example",
                "content": content,
                "source_name": source_name,
                "created_at": str(item.get("created_at", "")).strip(),
            }
        )
    return cleaned


def save_blog_style_example(
    title: str,
    content: str,
    source_name: str = "",
    max_items: int = 80,
) -> tuple[int, bool]:
    cleaned_title = str(title or "").strip() or "Untitled example"
    cleaned_content = str(content or "").strip()
    cleaned_source_name = str(source_name or "").strip()
    if not cleaned_content:
        return len(get_blog_style_examples()), False

    config_data = load_app_config()
    existing = config_data.get("blog_style_examples", [])
    if not isinstance(existing, list):
        existing = []
    existing = [item for item in existing if isinstance(item, dict)]

    for item in existing:
        existing_content = str(item.get("content", "")).strip()
        if is_near_duplicate_style_example(cleaned_content, existing_content):
            return len(existing), False

    entry = {
        "title": cleaned_title,
        "content": cleaned_content,
        "source_name": cleaned_source_name,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    existing.insert(0, entry)
    config_data["blog_style_examples"] = existing[:max_items]
    save_app_config(config_data)
    return len(config_data["blog_style_examples"]), True


def delete_blog_style_example(index: int) -> bool:
    config_data = load_app_config()
    examples = config_data.get("blog_style_examples", [])
    if not isinstance(examples, list):
        return False
    examples = [item for item in examples if isinstance(item, dict)]
    if index < 0 or index >= len(examples):
        return False
    examples.pop(index)
    config_data["blog_style_examples"] = examples
    save_app_config(config_data)
    return True


def normalize_style_example_text(text: str) -> str:
    normalized = re.sub(r"\s+", " ", str(text or "").strip().lower())
    normalized = re.sub(r"[^a-z0-9\s'-]", "", normalized)
    return normalized.strip()


def is_near_duplicate_style_example(candidate: str, existing: str) -> bool:
    a = normalize_style_example_text(candidate)
    b = normalize_style_example_text(existing)
    if not a or not b:
        return False
    if a == b:
        return True

    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    if len(shorter) >= 160 and shorter in longer:
        return True

    ratio = SequenceMatcher(None, a[:5000], b[:5000]).ratio()
    if ratio >= 0.9:
        return True

    tokens_a = set(re.findall(r"\b[a-z0-9'-]{4,}\b", a))
    tokens_b = set(re.findall(r"\b[a-z0-9'-]{4,}\b", b))
    if not tokens_a or not tokens_b:
        return False
    overlap = len(tokens_a & tokens_b) / max(1, len(tokens_a | tokens_b))
    return ratio >= 0.82 and overlap >= 0.8


def extract_text_from_docx_bytes(data: bytes) -> str:
    try:
        with zipfile.ZipFile(BytesIO(data)) as zf:
            if "word/document.xml" not in zf.namelist():
                return ""
            xml_bytes = zf.read("word/document.xml")
    except Exception:
        return ""

    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return ""

    paragraphs: list[str] = []
    for node in root.iter():
        if node.tag.endswith("}p"):
            parts: list[str] = []
            for inner in node.iter():
                if inner.tag.endswith("}t") and inner.text:
                    parts.append(inner.text)
            paragraph = "".join(parts).strip()
            if paragraph:
                paragraphs.append(paragraph)
    return "\n\n".join(paragraphs)


def build_style_profile(
    style_examples: list[str] | None,
    style_strength: str = "Balanced",
    style_example_names: list[str] | None = None,
) -> dict[str, object]:
    texts = [str(t).strip() for t in (style_examples or []) if str(t).strip()]
    names = [str(n).strip() for n in (style_example_names or []) if str(n).strip()]
    if not texts:
        return {
            "avg_sentence_words": 18,
            "starter_phrases": [],
            "signature_terms": [],
        }

    name_hint = "\n".join([f"Reference title: {n}" for n in names])
    combined = "\n".join(texts + ([name_hint] if name_hint else []))
    sentences = [s.strip() for s in re.split(r"(?<=[.!?])\s+", combined) if s.strip()]

    sentence_lengths = [len(re.findall(r"\b[\w'-]+\b", sentence)) for sentence in sentences if sentence]
    avg_sentence_words = int(round(sum(sentence_lengths) / len(sentence_lengths))) if sentence_lengths else 18

    starters: list[str] = []
    for sentence in sentences:
        words = re.findall(r"[A-Za-z][A-Za-z'-]*", sentence)
        if len(words) >= 2:
            starters.append(f"{words[0]} {words[1]}")
    starter_counts = Counter(starters)
    starter_phrases = [p for p, _ in starter_counts.most_common(6)]

    stop_words = {
        "the", "and", "for", "that", "with", "this", "from", "your", "into", "have", "will", "what", "when",
        "where", "which", "should", "could", "about", "their", "they", "them", "through", "there", "while",
        "because", "very", "more", "most", "only", "also", "than", "then", "just", "been", "being", "into",
    }
    words = [
        w.lower()
        for w in re.findall(r"\b[A-Za-z][A-Za-z'-]{4,}\b", combined)
        if w.lower() not in stop_words
    ]
    signature_terms = [term for term, _ in Counter(words).most_common(8)]

    bounded_avg = max(12, min(26, avg_sentence_words))
    strength = str(style_strength or "Balanced").strip().title()
    if strength not in {"Low", "Balanced", "High"}:
        strength = "Balanced"

    if strength == "Low":
        adjusted_avg = int(round((18 * 0.75) + (bounded_avg * 0.25)))
        starters_out = starter_phrases[:2]
        terms_out = signature_terms[:2]
    elif strength == "High":
        adjusted_avg = bounded_avg
        starters_out = starter_phrases[:6]
        terms_out = signature_terms[:8]
    else:
        adjusted_avg = int(round((18 * 0.5) + (bounded_avg * 0.5)))
        starters_out = starter_phrases[:4]
        terms_out = signature_terms[:4]

    return {
        "avg_sentence_words": max(12, min(26, adjusted_avg)),
        "starter_phrases": starters_out,
        "signature_terms": terms_out,
    }


def render_char_counter(label: str, text: str, limit: int) -> None:
    count = len(text or "")
    remaining = limit - count
    if remaining >= 0:
        color = "#166534" if remaining > max(int(limit * 0.2), 20) else "#9a3412"
        status = f"{remaining:,} left"
    else:
        color = "#b91c1c"
        status = f"{abs(remaining):,} over"
    st.markdown(
        (
            "<div style=\"display:inline-block;padding:0.2rem 0.55rem;border-radius:999px;"
            "border:1px solid #cbd5e1;background:#f8fafc;font-size:0.78rem;font-weight:600;"
            f"color:{color};margin-top:0.2rem;\">{html.escape(label)}: {count:,}/{limit:,} ({status})</div>"
        ),
        unsafe_allow_html=True,
    )


def build_draft_export_bundle(entry: dict, source_sales: pd.DataFrame) -> bytes:
    blog_text = str(entry.get("blog_text", "")).strip()
    cta = str(entry.get("cta", "book a short strategy call with me")).strip() or "book a short strategy call with me"
    target_mode = str(entry.get("target_mode", "Lead Staff (Client Manager)"))
    selected_person = str(entry.get("selected_person", "")).strip()

    recipients = source_sales.copy()
    if target_mode in recipients.columns and selected_person:
        recipients = recipients[recipients[target_mode].astype(str).str.strip() == selected_person].copy()

    recipient_columns = [
        c
        for c in ["Prospect Name", "Business Name", "Email", "Contact Phone #", "Partner", "Lead Staff (Client Manager)"]
        if c in recipients.columns
    ]
    recipients = recipients[recipient_columns].drop_duplicates() if recipient_columns else pd.DataFrame()

    messages = build_outreach_messages(blog_text, cta)

    draft_row = {
        "created_at": entry.get("created_at", ""),
        "title": entry.get("title", ""),
        "topic": entry.get("topic", ""),
        "audience": entry.get("audience", ""),
        "objective": entry.get("objective", ""),
        "tone": entry.get("tone", ""),
        "length": entry.get("length", ""),
        "cta": cta,
        "target_mode": target_mode,
        "selected_person": selected_person,
        "blog_text": blog_text,
    }
    draft_df = pd.DataFrame([draft_row])

    channels = ["email", "sms", "facebook", "instagram", "linkedin"]
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("draft_summary.csv", draft_df.to_csv(index=False))
        zf.writestr("recipients_snapshot.csv", recipients.to_csv(index=False))

        for channel in channels:
            channel_df = build_campaign_csv(recipients, channel, messages)
            zf.writestr(f"{channel}_campaign.csv", channel_df.to_csv(index=False))

    return zip_buffer.getvalue()


def render_blog_page(source_sales: pd.DataFrame, is_manager: bool) -> None:
    section_banner("Blog", "Create AI-assisted blog content, then prepare outreach messages for your clients and prospects.")

    st.markdown(
        """
        <style>
        .blog-hero {
            padding: 0.95rem 1rem;
            border-radius: 14px;
            border: 1px solid #f59e0b;
            background: linear-gradient(120deg, #fffbeb 0%, #fff7ed 55%, #f0fdf4 100%);
            margin-bottom: 0.75rem;
        }
        .blog-panel {
            padding: 0.75rem 0.85rem;
            border-radius: 12px;
            border: 1px solid #dbeafe;
            background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
            margin: 0.5rem 0 0.8rem 0;
        }
        .blog-panel h4 {
            margin: 0 0 0.25rem 0;
            color: #0f172a;
            font-size: 1.03rem;
        }
        .blog-panel p {
            margin: 0;
            color: #334155;
            font-size: 0.9rem;
        }
        .blog-principle-chip {
            display: inline-block;
            padding: 0.28rem 0.62rem;
            border-radius: 999px;
            font-size: 0.82rem;
            font-weight: 800;
            margin: 0.1rem 0 0.45rem 0;
            border: 2px solid transparent;
        }
        .blog-principle-chip.p1 {
            background: #fff7ed;
            color: #9a3412;
            border-color: #f97316;
        }
        .blog-principle-chip.p2 {
            background: #fdf2f8;
            color: #9d174d;
            border-color: #ec4899;
        }
        .blog-principle-chip.p3 {
            background: #eff6ff;
            color: #1d4ed8;
            border-color: #3b82f6;
        }

        .st-key-blog_principle_1_title input,
        .st-key-blog_principle_1_detail_1 input,
        .st-key-blog_principle_1_detail_2 input,
        .st-key-blog_principle_1_detail_3 input {
            background: #fff7ed !important;
            border: 2px solid #f97316 !important;
        }
        .st-key-blog_principle_1_title label p,
        .st-key-blog_principle_1_detail_1 label p,
        .st-key-blog_principle_1_detail_2 label p,
        .st-key-blog_principle_1_detail_3 label p {
            color: #9a3412 !important;
            font-weight: 700 !important;
        }

        .st-key-blog_principle_2_title input,
        .st-key-blog_principle_2_detail_1 input,
        .st-key-blog_principle_2_detail_2 input,
        .st-key-blog_principle_2_detail_3 input {
            background: #fdf2f8 !important;
            border: 2px solid #ec4899 !important;
        }
        .st-key-blog_principle_2_title label p,
        .st-key-blog_principle_2_detail_1 label p,
        .st-key-blog_principle_2_detail_2 label p,
        .st-key-blog_principle_2_detail_3 label p {
            color: #9d174d !important;
            font-weight: 700 !important;
        }

        .st-key-blog_principle_3_title input,
        .st-key-blog_principle_3_detail_1 input,
        .st-key-blog_principle_3_detail_2 input,
        .st-key-blog_principle_3_detail_3 input {
            background: #eff6ff !important;
            border: 2px solid #3b82f6 !important;
        }
        .st-key-blog_principle_3_title label p,
        .st-key-blog_principle_3_detail_1 label p,
        .st-key-blog_principle_3_detail_2 label p,
        .st-key-blog_principle_3_detail_3 label p {
            color: #1d4ed8 !important;
            font-weight: 700 !important;
        }

        #blog-outreach-tabs-anchor + div[data-testid="stTabs"] button[role="tab"] {
            border-width: 2px !important;
            border-style: solid !important;
            border-radius: 999px !important;
            font-weight: 700 !important;
        }
        #blog-outreach-tabs-anchor + div[data-testid="stTabs"] button[role="tab"]:nth-of-type(1) {
            background: #fff7ed !important;
            border-color: #f97316 !important;
            color: #9a3412 !important;
        }
        #blog-outreach-tabs-anchor + div[data-testid="stTabs"] button[role="tab"]:nth-of-type(2) {
            background: #fdf2f8 !important;
            border-color: #ec4899 !important;
            color: #9d174d !important;
        }
        #blog-outreach-tabs-anchor + div[data-testid="stTabs"] button[role="tab"]:nth-of-type(3) {
            background: #eff6ff !important;
            border-color: #3b82f6 !important;
            color: #1d4ed8 !important;
        }
        #blog-outreach-tabs-anchor + div[data-testid="stTabs"] button[role="tab"]:nth-of-type(4) {
            background: #fff1f2 !important;
            border-color: #f43f5e !important;
            color: #9f1239 !important;
        }
        #blog-outreach-tabs-anchor + div[data-testid="stTabs"] button[role="tab"]:nth-of-type(5) {
            background: #ecfdf3 !important;
            border-color: #22c55e !important;
            color: #166534 !important;
        }
        #blog-outreach-tabs-anchor + div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
            box-shadow: 0 0 0 3px rgba(15, 23, 42, 0.08) !important;
            transform: translateY(-1px);
        }
        </style>
        <div class="blog-hero">
            <div style="font-weight:800;font-size:1.05rem;color:#7c2d12;">Campaign Studio</div>
            <div style="font-size:0.9rem;color:#78350f;">Draft once, repurpose across channels, and keep your best posts in reusable history.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    jump_col, _ = st.columns([1.3, 4.7])
    with jump_col:
        if st.button("Go to Pipeline", use_container_width=True):
            st.session_state["current_page"] = "Pipeline"
            st.rerun()

    is_authorized = is_manager or st.checkbox(
        "I confirm I am authorized to send outreach communications to selected contacts.",
        value=False,
    )
    if not is_authorized:
        st.info("Authorization is required before creating or preparing outreach campaigns.")
        return

    st.caption("Recipient targeting is based on the full Sales Activity dataset and filtered by your selected Partner or Team Member identity.")

    mode_options = ["Partner", "Lead Staff (Client Manager)"]
    current_mode = st.session_state.get("blog_target_mode", mode_options[0])
    if current_mode not in mode_options:
        current_mode = mode_options[0]
    target_mode = st.radio("Target contacts by", mode_options, horizontal=True, key="blog_target_mode")
    mode_col = target_mode

    people_options = sorted(
        [
            str(v).strip()
            for v in source_sales.get(mode_col, pd.Series(dtype=str)).dropna().unique().tolist()
            if str(v).strip()
        ]
    )
    if not people_options:
        st.warning(f"No values found in column '{mode_col}' for targeting.")
        return

    preferred_person = st.session_state.get("blog_selected_person")
    selected_person_index = people_options.index(preferred_person) if preferred_person in people_options else 0
    selected_person = st.selectbox(
        "Select your name",
        options=people_options,
        index=selected_person_index,
        key="blog_selected_person",
    )
    recipient_df = source_sales[source_sales.get(mode_col, pd.Series(dtype=str)).astype(str).str.strip() == selected_person].copy()
    if recipient_df.empty:
        st.warning("No client/prospect rows found for the selected name.")
        return

    recipient_columns = [c for c in ["Prospect Name", "Business Name", "Email", "Contact Phone #", mode_col] if c in recipient_df.columns]
    recipients = recipient_df[recipient_columns].copy().drop_duplicates()

    k1, k2, k3 = st.columns(3)
    k1.metric("Recipients", f"{len(recipients):,}")
    k2.metric("Emails", f"{int(recipients.get('Email', pd.Series(dtype=str)).astype(str).str.strip().ne('').sum()):,}")
    k3.metric("Phone numbers", f"{int(recipients.get('Contact Phone #', pd.Series(dtype=str)).astype(str).str.strip().ne('').sum()):,}")

    st.dataframe(recipients.head(20), use_container_width=True, height=240)

    st.markdown('<div class="blog-panel"><h4>AI Blog Composer</h4><p>Shape the post using your voice, then keep successful drafts in history for future campaigns.</p></div>', unsafe_allow_html=True)

    history = get_blog_draft_history()
    with st.expander("Saved Draft History", expanded=False):
        if not history:
            st.caption("No saved drafts yet.")
        else:
            history_labels = []
            history_details = []
            for item in history:
                created = str(item.get("created_at", "")).strip() or "Unknown time"
                title = str(item.get("title", "Untitled")).strip() or "Untitled"
                owner = str(item.get("selected_person", "")).strip()
                style_strength_label = str(item.get("style_strength", "Balanced")).strip() or "Balanced"
                conf = item.get("style_conf_thresholds", {})
                moderate_min = 120
                strong_min = 300
                if isinstance(conf, dict):
                    try:
                        moderate_min = int(conf.get("moderate_min", 120))
                        strong_min = int(conf.get("strong_min", 300))
                    except (TypeError, ValueError):
                        moderate_min = 120
                        strong_min = 300
                if strong_min <= moderate_min:
                    strong_min = moderate_min + 10
                suffix = f" | {owner}" if owner else ""
                history_labels.append(f"{created} | {title[:80]}{suffix}")
                history_details.append(
                    {
                        "style_strength": style_strength_label,
                        "moderate_min": moderate_min,
                        "strong_min": strong_min,
                    }
                )

            selected_history_index = st.selectbox(
                "Select saved draft",
                options=list(range(len(history_labels))),
                format_func=lambda idx: history_labels[idx],
                key="blog_history_index",
            )
            if 0 <= int(selected_history_index) < len(history_details):
                detail = history_details[int(selected_history_index)]
                strength_value = str(detail.get("style_strength", "Balanced")).strip().title()
                strength_color = {
                    "Low": {"bg": "#fffbeb", "border": "#f59e0b", "text": "#92400e"},
                    "Balanced": {"bg": "#fefce8", "border": "#facc15", "text": "#854d0e"},
                    "High": {"bg": "#ecfdf3", "border": "#22c55e", "text": "#166534"},
                }.get(strength_value, {"bg": "#f8fafc", "border": "#cbd5e1", "text": "#334155"})
                moderate_value = int(detail.get("moderate_min", 120) or 120)
                strong_value = int(detail.get("strong_min", 300) or 300)
                st.markdown(
                    (
                        "<div style=\"display:flex;gap:0.35rem;flex-wrap:wrap;margin:0.15rem 0 0.4rem 0;\">"
                        f"<span style=\"display:inline-block;padding:0.2rem 0.55rem;border-radius:999px;border:1px solid {strength_color['border']};"
                        f"background:{strength_color['bg']};color:{strength_color['text']};font-size:0.78rem;font-weight:700;\">"
                        f"Strength: {html.escape(strength_value)}</span>"
                        "<span style=\"display:inline-block;padding:0.2rem 0.55rem;border-radius:999px;border:1px solid #facc15;"
                        "background:#fefce8;color:#854d0e;font-size:0.78rem;font-weight:700;\">"
                        f"Moderate: {moderate_value}</span>"
                        "<span style=\"display:inline-block;padding:0.2rem 0.55rem;border-radius:999px;border:1px solid #22c55e;"
                        "background:#ecfdf3;color:#166534;font-size:0.78rem;font-weight:700;\">"
                        f"Strong: {strong_value}</span>"
                        "</div>"
                    ),
                    unsafe_allow_html=True,
                )
            h_col1, h_col2 = st.columns(2)
            with h_col1:
                if st.button("Load Selected Draft", use_container_width=True):
                    chosen = history[selected_history_index]
                    st.session_state["blog_topic"] = str(chosen.get("topic", ""))
                    st.session_state["blog_audience"] = str(chosen.get("audience", ""))
                    st.session_state["blog_objective"] = str(chosen.get("objective", ""))
                    st.session_state["blog_tone"] = str(chosen.get("tone", "Professional"))
                    st.session_state["blog_length"] = str(chosen.get("length", "Medium"))
                    st.session_state["blog_cta"] = str(chosen.get("cta", "book a short strategy call with me"))
                    st.session_state["blog_polish_level"] = str(chosen.get("polish_level", "Strong"))
                    st.session_state["blog_rewrite_intensity"] = str(chosen.get("rewrite_intensity", "Editorial"))
                    st.session_state["blog_style_strength"] = str(chosen.get("style_strength", "Balanced"))
                    saved_style_thresholds = chosen.get("style_conf_thresholds", {})
                    if isinstance(saved_style_thresholds, dict):
                        try:
                            moderate_saved = int(saved_style_thresholds.get("moderate_min", st.session_state.get("blog_style_conf_moderate_min", 120)))
                            strong_saved = int(saved_style_thresholds.get("strong_min", st.session_state.get("blog_style_conf_strong_min", 300)))
                            if strong_saved <= moderate_saved:
                                strong_saved = moderate_saved + 10
                            st.session_state["blog_style_conf_moderate_min"] = moderate_saved
                            st.session_state["blog_style_conf_strong_min"] = strong_saved
                        except (TypeError, ValueError):
                            pass
                    saved_ranges = chosen.get("length_ranges", {})
                    if isinstance(saved_ranges, dict):
                        for range_key, state_key in [
                            ("short_min", "blog_len_short_min"),
                            ("short_max", "blog_len_short_max"),
                            ("medium_min", "blog_len_medium_min"),
                            ("medium_max", "blog_len_medium_max"),
                            ("long_min", "blog_len_long_min"),
                            ("long_max", "blog_len_long_max"),
                        ]:
                            if range_key in saved_ranges:
                                try:
                                    st.session_state[state_key] = int(saved_ranges[range_key])
                                except (TypeError, ValueError):
                                    pass
                    chosen_principles = chosen.get("principles", [])
                    if isinstance(chosen_principles, list) and chosen_principles:
                        for idx in range(1, 4):
                            p = chosen_principles[idx - 1] if idx - 1 < len(chosen_principles) else {}
                            p_title = str(p.get("title", "")).strip() if isinstance(p, dict) else ""
                            p_details = p.get("details", []) if isinstance(p, dict) else []
                            st.session_state[f"blog_principle_{idx}_title"] = p_title
                            for d_idx in range(1, 4):
                                d_val = ""
                                if isinstance(p_details, list) and d_idx - 1 < len(p_details):
                                    d_val = str(p_details[d_idx - 1]).strip()
                                st.session_state[f"blog_principle_{idx}_detail_{d_idx}"] = d_val
                    else:
                        legacy_points = [line.strip() for line in str(chosen.get("points", "")).splitlines() if line.strip()]
                        for idx in range(1, 4):
                            st.session_state[f"blog_principle_{idx}_title"] = legacy_points[idx - 1] if idx - 1 < len(legacy_points) else ""
                            for d_idx in range(1, 4):
                                st.session_state[f"blog_principle_{idx}_detail_{d_idx}"] = ""
                    loaded_outline = str(chosen.get("outline_text", chosen.get("blog_text", "")))
                    loaded_final = str(chosen.get("final_post_text", chosen.get("blog_text", "")))
                    st.session_state["blog_draft_text"] = loaded_outline
                    st.session_state["blog_draft_editor"] = loaded_outline
                    st.session_state["blog_final_text"] = loaded_final
                    st.session_state["blog_final_editor"] = loaded_final
                    saved_style_titles = chosen.get("style_example_titles", [])
                    if isinstance(saved_style_titles, list):
                        st.session_state["blog_style_selected_titles_pending"] = [
                            str(t).strip() for t in saved_style_titles if str(t).strip()
                        ]
                    loaded_mode = str(chosen.get("target_mode", "Partner"))
                    if loaded_mode in mode_options:
                        st.session_state["blog_target_mode"] = loaded_mode
                    loaded_person = str(chosen.get("selected_person", "")).strip()
                    if loaded_person in people_options:
                        st.session_state["blog_selected_person"] = loaded_person
                    st.success("Draft loaded from history.")
                    st.rerun()
            with h_col2:
                if st.button("Delete Selected Draft", use_container_width=True):
                    if delete_blog_draft_from_history(selected_history_index):
                        st.success("Draft removed from history.")
                        st.rerun()

            selected_history_entry = history[selected_history_index]
            bundle_bytes = build_draft_export_bundle(selected_history_entry, source_sales)
            raw_name = str(selected_history_entry.get("title", "blog_draft")).strip().lower()
            safe_name = re.sub(r"[^a-z0-9]+", "_", raw_name).strip("_") or "blog_draft"
            st.download_button(
                "Export Selected Draft Bundle (ZIP)",
                data=bundle_bytes,
                file_name=f"{safe_name}_bundle.zip",
                mime="application/zip",
                use_container_width=True,
                help="Includes draft summary, recipients snapshot, and all channel campaign CSV files.",
            )

    style_examples = get_blog_style_examples()
    style_labels = [
        (
            f"{idx + 1}. {str(item.get('title', 'Untitled example')).strip() or 'Untitled example'}"
            f" ({count_words(str(item.get('content', ''))):,} words)"
            + (
                f" - {str(item.get('source_name', '')).strip()}"
                if str(item.get('source_name', '')).strip()
                else ""
            )
        )
        for idx, item in enumerate(style_examples)
    ]
    if "blog_style_selected_labels" not in st.session_state:
        st.session_state["blog_style_selected_labels"] = style_labels[: min(3, len(style_labels))]
    else:
        st.session_state["blog_style_selected_labels"] = [
            label for label in st.session_state.get("blog_style_selected_labels", []) if label in style_labels
        ]

    pending_titles = st.session_state.pop("blog_style_selected_titles_pending", None)
    if isinstance(pending_titles, list) and pending_titles:
        selected_labels_from_titles: list[str] = []
        for idx, item in enumerate(style_examples):
            item_title = str(item.get("title", "")).strip()
            if item_title and item_title in pending_titles and idx < len(style_labels):
                selected_labels_from_titles.append(style_labels[idx])
        st.session_state["blog_style_selected_labels"] = selected_labels_from_titles

    selected_style_titles: list[str] = []
    selected_style_contents: list[str] = []
    selected_style_sources: list[str] = []
    selected_style_entries: list[dict[str, str]] = []
    style_strength = str(st.session_state.get("blog_style_strength", "Balanced")).strip().title()
    if style_strength not in {"Low", "Balanced", "High"}:
        style_strength = "Balanced"
    st.session_state["blog_style_strength"] = style_strength
    if "blog_style_conf_moderate_min" not in st.session_state:
        st.session_state["blog_style_conf_moderate_min"] = 120
    if "blog_style_conf_strong_min" not in st.session_state:
        st.session_state["blog_style_conf_strong_min"] = 300

    with st.expander("Style Library (Example Upload)", expanded=False):
        st.caption("Add examples of your past writing. Select examples below to steer final-post phrasing and cadence.")

        sx1, sx2 = st.columns(2)
        with sx1:
            example_title = st.text_input(
                "Article/Post name",
                key="blog_style_example_title",
                placeholder="e.g., Planning Ahead in Volatile Markets",
            )
            example_source = st.text_input(
                "Source or publication (optional)",
                key="blog_style_example_source",
                placeholder="e.g., LinkedIn, Newsletter, Website",
            )
        with sx2:
            uploaded_examples = st.file_uploader(
                "Upload examples (.txt, .md, .docx)",
                type=["txt", "md", "docx"],
                accept_multiple_files=True,
                key="blog_style_example_files",
            )
            upload_source = st.text_input(
                "Source for uploaded files (optional)",
                key="blog_style_upload_source",
                placeholder="e.g., Client newsletter archive",
            )

        style_strength = st.select_slider(
            "Style adaptation strength",
            options=["Low", "Balanced", "High"],
            value=st.session_state.get("blog_style_strength", "Balanced"),
            key="blog_style_strength",
            help="Low keeps your current generator voice dominant, Balanced blends both, High follows your uploaded examples more strongly.",
        )

        st.caption("Confidence thresholds: tune the word-count breakpoints for low, moderate, and strong style-signal badges.")
        ct1, ct2 = st.columns(2)
        with ct1:
            st.number_input(
                "Moderate starts at",
                min_value=20,
                max_value=10000,
                step=10,
                key="blog_style_conf_moderate_min",
            )
        with ct2:
            st.number_input(
                "Strong starts at",
                min_value=30,
                max_value=10000,
                step=10,
                key="blog_style_conf_strong_min",
            )

        moderate_min = int(st.session_state.get("blog_style_conf_moderate_min", 120))
        strong_min = int(st.session_state.get("blog_style_conf_strong_min", 300))
        if strong_min <= moderate_min:
            strong_min = moderate_min + 10
            st.session_state["blog_style_conf_strong_min"] = strong_min
            st.info("Adjusted strong threshold so it remains above moderate threshold.")

        example_text = st.text_area(
            "Paste an example post",
            key="blog_style_example_text",
            height=170,
            placeholder="Paste a past blog/article here to help model your writing style.",
        )

        sb1, sb2 = st.columns(2)
        with sb1:
            if st.button("Save Pasted Example", use_container_width=True):
                if not str(example_text).strip():
                    st.warning("Paste example text before saving.")
                else:
                    total_examples, saved = save_blog_style_example(
                        example_title,
                        example_text,
                        source_name=example_source,
                    )
                    if saved:
                        st.success(f"Example saved. Style library now has {total_examples} item(s).")
                    else:
                        st.info("A very similar example already exists, so this one was skipped.")
                    st.rerun()
        with sb2:
            if st.button("Save Uploaded Files", use_container_width=True):
                saved_count = 0
                skipped_count = 0
                for file in uploaded_examples or []:
                    file_name = str(getattr(file, "name", ""))
                    suffix = Path(file_name).suffix.lower()
                    try:
                        raw_data = file.getvalue()
                    except Exception:
                        raw_data = b""

                    content = ""
                    if raw_data:
                        if suffix == ".docx":
                            content = extract_text_from_docx_bytes(raw_data)
                        else:
                            content = raw_data.decode("utf-8", errors="ignore")

                    if str(content).strip():
                        _, saved = save_blog_style_example(
                            Path(file_name).stem,
                            content,
                            source_name=upload_source,
                        )
                        if saved:
                            saved_count += 1
                        else:
                            skipped_count += 1
                if saved_count and skipped_count:
                    st.success(f"Saved {saved_count} uploaded example(s); skipped {skipped_count} near-duplicate file(s).")
                    st.rerun()
                elif saved_count:
                    st.success(f"Saved {saved_count} uploaded example(s).")
                    st.rerun()
                elif skipped_count:
                    st.info("All uploaded files were already represented by very similar saved examples.")
                else:
                    st.warning("No readable text found in uploaded files.")

        if not style_examples:
            st.caption("No style examples saved yet.")
        else:
            st.multiselect(
                "Examples to apply for style adaptation",
                options=style_labels,
                key="blog_style_selected_labels",
                help="Selected examples influence sentence rhythm and recurring phrasing in the final post.",
            )

            selected_labels_now = set(st.session_state.get("blog_style_selected_labels", []))
            selected_for_preview: list[str] = []
            selected_titles_preview: list[str] = []
            for idx, label in enumerate(style_labels):
                if label in selected_labels_now and idx < len(style_examples):
                    selected_for_preview.append(str(style_examples[idx].get("content", "")).strip())
                    selected_titles_preview.append(str(style_examples[idx].get("title", "")).strip())

            st.markdown("##### Live style signals")
            if selected_for_preview:
                profile_preview = build_style_profile(
                    selected_for_preview,
                    style_strength=style_strength,
                    style_example_names=selected_titles_preview,
                )
                total_style_words = sum(count_words(text) for text in selected_for_preview)
                moderate_min = int(st.session_state.get("blog_style_conf_moderate_min", 120))
                strong_min = int(st.session_state.get("blog_style_conf_strong_min", 300))
                if strong_min <= moderate_min:
                    strong_min = moderate_min + 10
                m1, m2, m3 = st.columns(3)
                with m1:
                    st.metric("Avg sentence length", f"{int(profile_preview.get('avg_sentence_words', 18))} words")
                with m2:
                    starters = [str(v).strip() for v in profile_preview.get("starter_phrases", []) if str(v).strip()]
                    st.caption("Common sentence starts")
                    st.write(", ".join(starters[:4]) if starters else "No strong starter pattern detected yet.")
                with m3:
                    terms = [str(v).strip() for v in profile_preview.get("signature_terms", []) if str(v).strip()]
                    st.caption("Recurring terms")
                    st.write(", ".join(terms[:6]) if terms else "No recurring term pattern detected yet.")
                if total_style_words < moderate_min:
                    st.markdown(
                        (
                            "<div style=\"display:inline-block;padding:0.25rem 0.6rem;border-radius:999px;"
                            "border:1px solid #f59e0b;background:#fffbeb;color:#92400e;font-size:0.78rem;font-weight:700;\">"
                            f"Low style-signal confidence: only {total_style_words} total words selected. Add more example text to reach at least {moderate_min} words."
                            "</div>"
                        ),
                        unsafe_allow_html=True,
                    )
                elif total_style_words < strong_min:
                    st.markdown(
                        (
                            "<div style=\"display:inline-block;padding:0.25rem 0.6rem;border-radius:999px;"
                            "border:1px solid #facc15;background:#fefce8;color:#854d0e;font-size:0.78rem;font-weight:700;\">"
                            f"Moderate style-signal confidence: {total_style_words} total words selected. Add more text to reach {strong_min} words for strong confidence."
                            "</div>"
                        ),
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        (
                            "<div style=\"display:inline-block;padding:0.25rem 0.6rem;border-radius:999px;"
                            "border:1px solid #22c55e;background:#ecfdf3;color:#166534;font-size:0.78rem;font-weight:700;\">"
                            f"Strong style-signal confidence: {total_style_words} total words selected."
                            "</div>"
                        ),
                        unsafe_allow_html=True,
                    )
            else:
                st.caption("Select at least one example to preview detected style signals.")

            selected_preview_index = st.selectbox(
                "Preview saved example",
                options=list(range(len(style_examples))),
                format_func=lambda idx: style_labels[idx],
                key="blog_style_preview_index",
            )
            preview_item = style_examples[selected_preview_index]
            preview_text = str(preview_item.get("content", "")).strip()
            st.text_area("Preview", value=preview_text, height=170, disabled=True)

            if st.button("Delete Previewed Example", use_container_width=True):
                if delete_blog_style_example(selected_preview_index):
                    st.success("Style example deleted.")
                    st.rerun()

    for idx, label in enumerate(style_labels):
        if label in st.session_state.get("blog_style_selected_labels", []) and idx < len(style_examples):
            example_item = style_examples[idx]
            selected_title = str(example_item.get("title", "")).strip() or "Untitled example"
            selected_style_titles.append(selected_title)
            selected_style_contents.append(str(example_item.get("content", "")).strip())
            source_val = str(example_item.get("source_name", "")).strip()
            if source_val:
                selected_style_sources.append(source_val)
            selected_style_entries.append({"title": selected_title, "source": source_val})

    if selected_style_contents:
        source_hint = f" | Sources: {', '.join(selected_style_sources[:3])}" if selected_style_sources else ""
        st.caption(
            f"Style adaptation active from {len(selected_style_contents)} example(s) at {style_strength} strength.{source_hint}"
        )

    c1, c2 = st.columns(2)
    with c1:
        topic = st.text_input("Blog topic", value="Planning ahead in uncertain markets", key="blog_topic")
        audience = st.text_input("Audience", value="clients and prospects", key="blog_audience")
        objective = st.text_input("Objective", value="clarity and practical next steps", key="blog_objective")
    with c2:
        tone_options = ["Professional", "Friendly", "Confident", "Educational"]
        tone_default = st.session_state.get("blog_tone", "Professional")
        tone_index = tone_options.index(tone_default) if tone_default in tone_options else 0
        tone = st.selectbox("Tone", options=tone_options, index=tone_index, key="blog_tone")
        length_options = ["Short", "Medium", "Long"]
        length_default = st.session_state.get("blog_length", "Medium")
        length_index = length_options.index(length_default) if length_default in length_options else 1
        length = st.selectbox("Length", options=length_options, index=length_index, key="blog_length")
        cta = st.text_input("Call to action", value="book a short strategy call with me", key="blog_cta")

        if "blog_len_short_min" not in st.session_state:
            st.session_state["blog_len_short_min"] = 30
        if "blog_len_short_max" not in st.session_state:
            st.session_state["blog_len_short_max"] = 90
        if "blog_len_medium_min" not in st.session_state:
            st.session_state["blog_len_medium_min"] = 250
        if "blog_len_medium_max" not in st.session_state:
            st.session_state["blog_len_medium_max"] = 400
        if "blog_len_long_min" not in st.session_state:
            st.session_state["blog_len_long_min"] = 600
        if "blog_len_long_max" not in st.session_state:
            st.session_state["blog_len_long_max"] = 900

        with st.expander("Length ranges (words)", expanded=False):
            st.caption("Set your preferred min and max words for each length option.")
            lr1, lr2 = st.columns(2)
            with lr1:
                st.number_input("Short min", min_value=10, max_value=2000, step=10, key="blog_len_short_min")
                st.number_input("Medium min", min_value=10, max_value=3000, step=10, key="blog_len_medium_min")
                st.number_input("Long min", min_value=10, max_value=5000, step=10, key="blog_len_long_min")
            with lr2:
                st.number_input("Short max", min_value=10, max_value=2000, step=10, key="blog_len_short_max")
                st.number_input("Medium max", min_value=10, max_value=3000, step=10, key="blog_len_medium_max")
                st.number_input("Long max", min_value=10, max_value=5000, step=10, key="blog_len_long_max")

        length_ranges = {
            "Short": (
                min(int(st.session_state["blog_len_short_min"]), int(st.session_state["blog_len_short_max"])),
                max(int(st.session_state["blog_len_short_min"]), int(st.session_state["blog_len_short_max"])),
            ),
            "Medium": (
                min(int(st.session_state["blog_len_medium_min"]), int(st.session_state["blog_len_medium_max"])),
                max(int(st.session_state["blog_len_medium_min"]), int(st.session_state["blog_len_medium_max"])),
            ),
            "Long": (
                min(int(st.session_state["blog_len_long_min"]), int(st.session_state["blog_len_long_max"])),
                max(int(st.session_state["blog_len_long_min"]), int(st.session_state["blog_len_long_max"])),
            ),
        }
        selected_min_words, selected_max_words = length_ranges.get(length, (250, 400))
        st.caption(f"{length} target: {selected_min_words} to {selected_max_words} words")

    st.markdown("#### Three Principles")
    st.caption("Define up to 3 principles. Each principle can include up to 3 detail points to expand in the post.")

    default_principles = [
        {
            "title": "Market context",
            "details": [
                "What changed recently",
                "Why it matters to clients",
                "Most relevant signal to watch",
            ],
        },
        {
            "title": "Client implications",
            "details": [
                "Likely risks",
                "Likely opportunities",
                "Common mistake to avoid",
            ],
        },
        {
            "title": "Practical next steps",
            "details": [
                "Immediate action",
                "30-day follow-up",
                "When to seek tailored advice",
            ],
        },
    ]
    for idx in range(1, 4):
        p_key = f"blog_principle_{idx}_title"
        if p_key not in st.session_state:
            st.session_state[p_key] = default_principles[idx - 1]["title"]
        for d_idx in range(1, 4):
            d_key = f"blog_principle_{idx}_detail_{d_idx}"
            if d_key not in st.session_state:
                st.session_state[d_key] = default_principles[idx - 1]["details"][d_idx - 1]

    for idx in range(1, 4):
        st.markdown(f'<div class="blog-principle-chip p{idx}">Principle {idx}</div>', unsafe_allow_html=True)
        st.text_input(
            f"Principle {idx} title",
            key=f"blog_principle_{idx}_title",
            placeholder="Enter core principle",
        )
        d_col1, d_col2, d_col3 = st.columns(3)
        with d_col1:
            st.text_input(
                f"Detail 1 (P{idx})",
                key=f"blog_principle_{idx}_detail_1",
                placeholder="Detail point 1",
            )
        with d_col2:
            st.text_input(
                f"Detail 2 (P{idx})",
                key=f"blog_principle_{idx}_detail_2",
                placeholder="Detail point 2",
            )
        with d_col3:
            st.text_input(
                f"Detail 3 (P{idx})",
                key=f"blog_principle_{idx}_detail_3",
                placeholder="Detail point 3",
            )

    principles_input: list[dict[str, object]] = []
    for idx in range(1, 4):
        p_title = str(st.session_state.get(f"blog_principle_{idx}_title", "")).strip()
        p_details = []
        for d_idx in range(1, 4):
            detail_value = str(st.session_state.get(f"blog_principle_{idx}_detail_{d_idx}", "")).strip()
            if detail_value:
                p_details.append(detail_value)
        if p_title or p_details:
            principles_input.append({"title": p_title, "details": p_details})

    if "blog_final_text" not in st.session_state:
        st.session_state["blog_final_text"] = ""
    if "blog_final_editor" not in st.session_state:
        st.session_state["blog_final_editor"] = st.session_state.get("blog_final_text", "")

    g_col1, g_col2 = st.columns([1.2, 1])
    with g_col1:
        generate_clicked = st.button("Generate Draft Outline", type="primary")
    with g_col2:
        save_clicked = st.button("Save Draft to History")

    if generate_clicked:
        st.session_state["blog_draft_text"] = build_blog_draft(
            topic,
            audience,
            objective,
            tone,
            length,
            principles_input,
            cta,
            word_range=(selected_min_words, selected_max_words),
        )
        st.session_state["blog_draft_editor"] = st.session_state["blog_draft_text"]

    if "blog_draft_editor_pending" in st.session_state:
        st.session_state["blog_draft_editor"] = st.session_state.pop("blog_draft_editor_pending")
    if "blog_final_editor_pending" in st.session_state:
        st.session_state["blog_final_editor"] = st.session_state.pop("blog_final_editor_pending")

    blog_text = st.text_area(
        "Draft outline",
        value=st.session_state.get("blog_draft_text", ""),
        height=360,
        key="blog_draft_editor",
    )

    st.markdown("#### Final Blog")
    st.caption("Review and edit the draft outline above, then generate your polished final post from those edits.")
    polish_options = ["Standard", "Strong", "Premium"]
    polish_default = st.session_state.get("blog_polish_level", "Strong")
    polish_index = polish_options.index(polish_default) if polish_default in polish_options else 1
    polish_level = st.selectbox(
        "Writing polish level",
        options=polish_options,
        index=polish_index,
        key="blog_polish_level",
        help="Standard = simple clear prose, Strong = more refined writing, Premium = most polished narrative style.",
    )
    rewrite_options = ["Conservative", "Editorial", "Publish-Ready"]
    rewrite_default = st.session_state.get("blog_rewrite_intensity", "Editorial")
    rewrite_index = rewrite_options.index(rewrite_default) if rewrite_default in rewrite_options else 1
    rewrite_intensity = st.selectbox(
        "Rewrite intensity",
        options=rewrite_options,
        index=rewrite_index,
        key="blog_rewrite_intensity",
        help="Conservative keeps closer to your draft, Editorial balances clarity and flow, Publish-Ready applies strongest narrative rewrite.",
    )
    st.markdown("##### Active Reference Posts")
    if selected_style_entries:
        for ref in selected_style_entries:
            ref_title = str(ref.get("title", "")).strip()
            ref_source = str(ref.get("source", "")).strip()
            if ref_source:
                st.caption(f"- {ref_title} ({ref_source})")
            else:
                st.caption(f"- {ref_title}")
    else:
        st.caption("No reference posts selected. Select examples in Style Library to shape final writing.")

    final_post_clicked = st.button("Generate Final Post from Draft Outline", type="primary", use_container_width=True)
    if final_post_clicked:
        if not blog_text.strip():
            st.warning("Create a draft outline first, then edit it before generating the final post.")
        else:
            st.session_state["blog_final_text"] = build_final_blog_post_from_outline(
                outline_text=blog_text,
                topic=topic,
                audience=audience,
                objective=objective,
                tone=tone,
                cta=cta,
                word_range=(selected_min_words, selected_max_words),
                polish_level=polish_level,
                rewrite_intensity=rewrite_intensity,
                style_examples=selected_style_contents,
                style_strength=style_strength,
                style_example_names=selected_style_titles,
            )
            st.session_state["blog_final_editor_pending"] = st.session_state["blog_final_text"]
            st.rerun()

    final_blog_text = st.text_area(
        "Finished blog (ready to send)",
        value=st.session_state.get("blog_final_text", ""),
        height=360,
        key="blog_final_editor",
    )

    if save_clicked:
        if not blog_text.strip() and not final_blog_text.strip():
            st.warning("Generate or enter a draft before saving to history.")
        else:
            selected_post = final_blog_text.strip() or blog_text.strip()
            title_line = next((line.lstrip("# ").strip() for line in selected_post.splitlines() if line.strip().startswith("#")), topic)
            entry = {
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "title": title_line,
                "topic": topic,
                "audience": audience,
                "objective": objective,
                "tone": tone,
                "length": length,
                "cta": cta,
                "polish_level": st.session_state.get("blog_polish_level", "Strong"),
                "rewrite_intensity": st.session_state.get("blog_rewrite_intensity", "Editorial"),
                "length_ranges": {
                    "short_min": int(st.session_state.get("blog_len_short_min", 30)),
                    "short_max": int(st.session_state.get("blog_len_short_max", 90)),
                    "medium_min": int(st.session_state.get("blog_len_medium_min", 250)),
                    "medium_max": int(st.session_state.get("blog_len_medium_max", 400)),
                    "long_min": int(st.session_state.get("blog_len_long_min", 600)),
                    "long_max": int(st.session_state.get("blog_len_long_max", 900)),
                },
                "principles": principles_input,
                "points": "\n".join([str(p.get("title", "")).strip() for p in principles_input if str(p.get("title", "")).strip()]),
                "outline_text": blog_text,
                "final_post_text": final_blog_text,
                "blog_text": selected_post,
                "style_example_titles": selected_style_titles,
                "style_strength": style_strength,
                "style_conf_thresholds": {
                    "moderate_min": int(st.session_state.get("blog_style_conf_moderate_min", 120)),
                    "strong_min": int(st.session_state.get("blog_style_conf_strong_min", 300)),
                },
                "target_mode": target_mode,
                "selected_person": selected_person,
            }
            total_saved = save_blog_draft_to_history(entry)
            st.success(f"Draft saved. History now has {total_saved} item(s).")

    st.caption(f"Draft outline words: {count_words(blog_text):,} | Final blog words: {count_words(final_blog_text):,}")

    outreach_source = final_blog_text.strip() or blog_text.strip()
    messages = build_outreach_messages(outreach_source, cta) if outreach_source else None
    if not messages:
        st.info("Generate a blog draft to prepare outreach messages.")
        return

    st.markdown('<div class="blog-panel"><h4>Outreach Studio</h4><p>Each channel has a practical length target. Use counters to keep posts clean and delivery-friendly.</p></div>', unsafe_allow_html=True)
    st.markdown('<div id="blog-outreach-tabs-anchor"></div>', unsafe_allow_html=True)
    t1, t2, t3, t4, t5 = st.tabs(["Email", "Text", "Facebook", "Instagram", "LinkedIn"])

    with t1:
        email_subject = st.text_input("Email subject", value=messages["email_subject"], key="email_subject_preview")
        render_char_counter("Subject", email_subject, 78)
        email_body = st.text_area("Email body", value=messages["email_body"], height=180, key="email_body_preview")
        render_char_counter("Email body", email_body, 2000)
        email_csv = build_campaign_csv(recipients, "email", messages)
        st.download_button(
            "Download email campaign CSV",
            data=email_csv.to_csv(index=False).encode("utf-8"),
            file_name=f"email_campaign_{selected_person.replace(' ', '_')}.csv",
            mime="text/csv",
        )

    with t2:
        sms_text = st.text_area("SMS message", value=messages["sms"], height=140, key="sms_preview")
        render_char_counter("SMS", sms_text, 160)
        sms_csv = build_campaign_csv(recipients, "sms", messages)
        st.download_button(
            "Download text campaign CSV",
            data=sms_csv.to_csv(index=False).encode("utf-8"),
            file_name=f"sms_campaign_{selected_person.replace(' ', '_')}.csv",
            mime="text/csv",
        )

    with t3:
        facebook_post = st.text_area("Facebook post", value=messages["facebook"], height=140, key="facebook_preview")
        render_char_counter("Facebook", facebook_post, 63206)
        facebook_csv = build_campaign_csv(recipients, "facebook", messages)
        st.download_button(
            "Download Facebook post draft",
            data=facebook_csv.to_csv(index=False).encode("utf-8"),
            file_name="facebook_post_draft.csv",
            mime="text/csv",
        )

    with t4:
        insta_post = st.text_area("Instagram post", value=messages["instagram"], height=140, key="instagram_preview")
        render_char_counter("Instagram", insta_post, 2200)
        insta_csv = build_campaign_csv(recipients, "instagram", messages)
        st.download_button(
            "Download Instagram post draft",
            data=insta_csv.to_csv(index=False).encode("utf-8"),
            file_name="instagram_post_draft.csv",
            mime="text/csv",
        )

    with t5:
        linkedin_post = st.text_area("LinkedIn post", value=messages["linkedin"], height=140, key="linkedin_preview")
        render_char_counter("LinkedIn", linkedin_post, 3000)
        linkedin_csv = build_campaign_csv(recipients, "linkedin", messages)
        st.download_button(
            "Download LinkedIn post draft",
            data=linkedin_csv.to_csv(index=False).encode("utf-8"),
            file_name="linkedin_post_draft.csv",
            mime="text/csv",
        )


def render_stats_to_date_detail(stats_df: pd.DataFrame) -> None:
    if stats_df.empty:
        st.info("No Stats to Date detail rows were found in range B2:G66.")
        return

    # C and F are separator columns in the workbook layout, so they are intentionally hidden.
    ordered_cols = [col for col in ["B", "D", "E", "G", "Row"] if col in stats_df.columns]
    view_df = stats_df[ordered_cols].copy().rename(columns={"Row": "Excel Row"})

    numeric_cols = [col for col in ["E", "G"] if col in view_df.columns]
    for col in numeric_cols:
        view_df[col] = pd.to_numeric(view_df[col], errors="coerce")

    section_rows = {2, 7, 18, 28, 39, 44, 55, 60}

    def format_whole(value: object) -> str:
        if pd.isna(value):
            return ""
        try:
            return f"{float(value):,.0f}"
        except (TypeError, ValueError):
            return str(value)

    def format_percent_or_whole(value: object) -> str:
        if pd.isna(value):
            return ""
        try:
            v = float(value)
        except (TypeError, ValueError):
            return str(value)
        if 0 <= v <= 1:
            return f"{v * 100:.0f}%"
        return f"{v:,.0f}"

    def color_band(series: pd.Series, palette: list[str]) -> list[str]:
        numeric = pd.to_numeric(series, errors="coerce")
        valid = numeric.dropna()
        if valid.empty:
            return [""] * len(series)

        min_v = float(valid.min())
        max_v = float(valid.max())
        if max_v == min_v:
            return [f"background-color: {palette[-1]};" if pd.notna(v) else "" for v in numeric]

        out: list[str] = []
        for value in numeric:
            if pd.isna(value):
                out.append("")
                continue
            ratio = (float(value) - min_v) / (max_v - min_v)
            if ratio < 0.25:
                color = palette[0]
            elif ratio < 0.5:
                color = palette[1]
            elif ratio < 0.75:
                color = palette[2]
            else:
                color = palette[3]
            out.append(f"background-color: {color};")
        return out

    def highlight_stats_rows(row: pd.Series) -> list[str]:
        excel_row = int(row.get("Excel Row", 0)) if str(row.get("Excel Row", "")).strip() else 0
        if excel_row in section_rows:
            return ["background-color: #dbeafe; color: #0f172a; font-weight: 700;"] * len(row)

        d_text = str(row.get("D", "")).strip().lower()
        if d_text == "(sub total)":
            return ["background-color: #fef3c7; color: #7c2d12; font-weight: 700;"] * len(row)

        return [""] * len(row)

    main_styler = (
        view_df.style
        .set_table_styles(
            [
                {"selector": "th", "props": "background-color: #0f4c81; color: white; font-weight: 700;"},
                {"selector": "td", "props": "border-bottom: 1px solid #e2e8f0;"},
            ]
        )
        .apply(lambda s: color_band(s, ["#fff7d6", "#ffe9a3", "#ffd166", "#f59e0b"]), subset=numeric_cols)
        .apply(highlight_stats_rows, axis=1)
        .format({
            "E": format_whole if "E" in view_df.columns else None,
            "G": format_percent_or_whole if "G" in view_df.columns else None,
        })
        .hide(subset=["Excel Row"], axis="columns")
    )

    def render_detail_cards(
        title: str,
        start_row: int,
        end_row: int,
        card_start: str,
        card_end: str,
        border_color: str,
        title_color: str,
        show_rate: bool = True,
    ) -> None:
        block = view_df[(view_df["Excel Row"] >= start_row) & (view_df["Excel Row"] <= end_row)]
        if block.empty:
            st.info(f"No rows found for B{start_row}:G{end_row}.")
            return

        cards_html: list[str] = []
        for _, row in block.iterrows():
            label = str(row.get("D", "")).strip() or str(row.get("B", "")).strip()
            if not label:
                continue
            if label.lower() == "(sub total)":
                continue
            if label.isdigit():
                continue

            category = str(row.get("B", "")).strip()
            count_val = format_whole(row.get("E", ""))
            rate_val = format_percent_or_whole(row.get("G", ""))
            esc_category = html.escape(category)
            esc_label = html.escape(label)
            esc_count = html.escape(str(count_val))
            esc_rate = html.escape(str(rate_val))

            category_html = ""
            if category and category.lower() != "(sub total)":
                category_html = (
                    f'<div style="font-size:0.75rem;color:#334155;margin-bottom:0.15rem;">{esc_category}</div>'
                )

            rate_html = ""
            if show_rate and esc_rate:
                rate_html = f'<span><strong>Rate:</strong> {esc_rate}</span>'

            card_html = (
                f'<div style="background:linear-gradient(135deg,{card_start} 0%,{card_end} 100%);'
                f'border:1px solid {border_color};border-radius:12px;padding:0.5rem 0.65rem;margin-bottom:0.4rem;">'
                f'{category_html}'
                f'<div style="font-weight:700;color:{title_color};margin-bottom:0.15rem;font-size:0.92rem;">{esc_label}</div>'
                f'<div style="display:flex;gap:0.7rem;color:#1f2937;font-size:0.86rem;">'
                f'<span><strong>Count:</strong> {esc_count}</span>'
                f'{rate_html}'
                f'</div></div>'
            )
            cards_html.append(card_html)

        if not cards_html:
            st.info(f"No detail rows to show for B{start_row}:G{end_row}.")
            return

        panel_html = (
            f'<div style="background:linear-gradient(180deg,#ffffff 0%,#f8fafc 100%);'
            f'border:2px solid {border_color};border-radius:14px;padding:0.55rem 0.6rem;'
            f'min-height:120px;">'
            + "".join(cards_html)
            + "</div>"
        )
        st.markdown(panel_html, unsafe_allow_html=True)

    st.markdown("**Detail Reports**")
    tabs = st.tabs([
        "Prospect Detail",
        "Pipeline Detail",
    ])

    with tabs[0]:
        p1, p2, p3 = st.columns(3)
        with p1:
            render_detail_cards("Relationship Type (B2:G3)", 2, 3, "#fee2e2", "#fecaca", "#fca5a5", "#991b1b")
        with p2:
            render_detail_cards("New Prospect Source (B7:G14)", 7, 14, "#fff7ed", "#ffedd5", "#fdba74", "#9a3412")
        with p3:
            render_detail_cards("Approach Style (B18:G26)", 18, 26, "#ecfeff", "#cffafe", "#67e8f9", "#0e7490")

    with tabs[1]:
        q1, q2, q3 = st.columns(3)
        with q1:
            render_detail_cards("Campaign Results (B28:G42)", 28, 42, "#dcfce7", "#bbf7d0", "#86efac", "#166534")
        with q2:
            render_detail_cards("Total Needs Stages (B44:G51)", 44, 51, "#ede9fe", "#ddd6fe", "#c4b5fd", "#4c1d95")
        with q3:
            render_detail_cards("Sales Process Status (B60:E65)", 60, 65, "#cffafe", "#99f6e4", "#5eead4", "#115e59", show_rate=False)


def data_tables(filtered_sales: pd.DataFrame, team: pd.DataFrame, coi: pd.DataFrame) -> None:
    tab1, tab2, tab3 = st.tabs(["Pipeline", "Team", "COI"])

    with tab1:
        display_cols = [
            "Prospect Name",
            "Business Name",
            "Lead Staff (Client Manager)",
            "Prospect Status",
            "Approach Date",
            "Secure Meeting",
            "Proposal Sent",
            "Job Secured",
            "Job Secured Value",
            "Comments / Next Move",
        ]
        cols = [c for c in display_cols if c in filtered_sales.columns]
        st.dataframe(filtered_sales[cols], use_container_width=True, height=420)

        csv_data = filtered_sales[cols].to_csv(index=False).encode("utf-8")
        st.download_button(
            label="Download filtered pipeline CSV",
            data=csv_data,
            file_name="sales_pipeline_filtered.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with tab2:
        st.dataframe(team, use_container_width=True, height=420)

    with tab3:
        st.dataframe(coi, use_container_width=True, height=420)


def render_lists_page(workbook_path: str) -> None:
    section_banner("Lists", "Manage dropdown options used across Pipeline and COI. Save writes directly to the workbook Lists sheet.")

    opts = st.session_state.get("dropdown_options", DROPDOWN_OPTIONS)
    updated: dict[str, dict[str, list[str]]] = {}

    for view_name, fields in opts.items():
        st.subheader(f"{view_name} Dropdowns")
        field_items = list(fields.items())
        # Render in two columns for compact layout
        col_count = 2
        cols = st.columns(col_count)
        updated[view_name] = {}
        for idx, (field, values) in enumerate(field_items):
            with cols[idx % col_count]:
                raw = st.text_area(
                    field,
                    value="\n".join(values),
                    height=140,
                    key=f"lists_{view_name}_{field}",
                    help="One option per line. Blank lines are ignored.",
                )
                updated[view_name][field] = [v.strip() for v in raw.splitlines() if v.strip()]
        st.divider()

    if st.button("Save List Changes", type="primary", use_container_width=False):
        try:
            updated_cells, truncated_fields = save_dropdown_options_to_lists(workbook_path, updated)
            st.session_state["dropdown_options"] = updated
            load_data.clear()
            if truncated_fields:
                st.warning(
                    "Some lists were longer than their workbook range and were trimmed: "
                    + ", ".join(truncated_fields)
                )
            st.success(f"Saved list changes to workbook ({updated_cells} cell updates).")
            st.rerun()
        except Exception as exc:
            st.error("Could not save list changes to workbook.")
            st.exception(exc)


def render_page(
    selected_page: str,
    source_sales: pd.DataFrame,
    filtered_sales: pd.DataFrame,
    team: pd.DataFrame,
    coi: pd.DataFrame,
    stats: pd.DataFrame,
    workbook_path: str,
    is_manager: bool = False,
) -> None:
    if selected_page == "Home":
        render_home_page(filtered_sales, team, coi)
    elif selected_page == "Dashboard":
        section_banner("Dashboard", "Headline numbers and charts now sit together in one view so you can scan the full picture without switching pages.")
        kpi_row(filtered_sales)
        charts(filtered_sales, coi)
        render_stats_to_date_detail(stats)
    elif selected_page == "Pipeline":
        section_banner("Pipeline", "Review active prospects and export the currently filtered pipeline in one place.")
        save_button_slot = st.empty()
        if "pipeline_new_rows_count" not in st.session_state:
            st.session_state["pipeline_new_rows_count"] = 0
        if "pipeline_import_rows" not in st.session_state:
            st.session_state["pipeline_import_rows"] = None
        pipeline_columns = get_sheet_columns_by_excel_range(
            workbook_path,
            "Sales Activity",
            "C",
            "AN",
            HEADER_MARKERS["Sales Activity"],
        )
        if not pipeline_columns:
            pipeline_columns = [
                "Prospect Name",
                "Business Name",
                "Lead Staff (Client Manager)",
                "Prospect Status",
                "Approach Date",
                "Secure Meeting",
                "Proposal Sent",
                "Job Secured",
                "Job Secured Value",
                "Comments / Next Move",
            ]
        pipeline_base_df = build_pipeline_base_df(filtered_sales, pipeline_columns)
        pipeline_base_df = append_imported_rows(pipeline_base_df, st.session_state.get("pipeline_import_rows"))
        pipeline_input_df = append_blank_rows(pipeline_base_df, int(st.session_state.get("pipeline_new_rows_count", 0)))
        pipeline_input_df = apply_new_row_defaults(pipeline_input_df, "Pipeline")

        with st.expander("Import client/prospect list", expanded=False):
            st.caption("Upload a CSV or Excel file. Matching columns will be appended as new Pipeline rows for review before save.")
            template_bytes = build_pipeline_import_template(
                pipeline_columns,
                st.session_state.get("dropdown_options", DROPDOWN_OPTIONS),
            )
            st.download_button(
                "Download blank Pipeline import template",
                data=template_bytes,
                file_name="pipeline_import_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=False,
            )
            uploaded_pipeline_file = st.file_uploader(
                "Import file",
                type=["csv", "xlsx", "xlsm", "xls"],
                key="pipeline_import_file",
                help="Supported formats: CSV, XLSX, XLSM, XLS.",
            )
            column_overrides: dict[str, str] = {}
            if uploaded_pipeline_file is not None:
                try:
                    preview_source_df = load_import_source(uploaded_pipeline_file)
                    preview_source_df = normalize_columns(preview_source_df)
                    st.caption("Optional: override any column mapping before import.")
                    override_cols = st.columns(2)
                    for idx, incoming_column in enumerate(preview_source_df.columns.tolist()):
                        with override_cols[idx % 2]:
                            selected_override = st.selectbox(
                                f"Map '{incoming_column}'",
                                options=["Auto-detect", "Ignore"] + pipeline_columns,
                                index=0,
                                key=f"pipeline_import_override_{canonicalize_column_name(uploaded_pipeline_file.name)}_{idx}",
                            )
                            if selected_override != "Auto-detect":
                                column_overrides[str(incoming_column)] = selected_override

                    preview_mapping_df, preview_matched_columns, preview_unmapped_columns = get_pipeline_import_mapping(
                        preview_source_df.columns.tolist(),
                        pipeline_columns,
                        column_overrides=column_overrides,
                    )
                    st.caption(f"Detected {len(preview_source_df)} rows in the selected file.")
                    if not preview_mapping_df.empty:
                        st.dataframe(preview_mapping_df, use_container_width=True, hide_index=True)
                    if preview_matched_columns:
                        st.caption("Columns ready to import: " + ", ".join(preview_matched_columns))
                    if preview_unmapped_columns:
                        st.warning("Columns that will be ignored: " + ", ".join(preview_unmapped_columns))
                except Exception as exc:
                    st.error("Could not preview the selected file.")
                    st.exception(exc)
            import_col, clear_import_col = st.columns([1.2, 1])
            with import_col:
                if st.button("Import Into Pipeline", use_container_width=True, disabled=uploaded_pipeline_file is None):
                    try:
                        imported_rows, matched_columns, unmapped_columns = parse_pipeline_import(
                            uploaded_pipeline_file,
                            pipeline_columns,
                            column_overrides=column_overrides,
                        )
                        if imported_rows.empty:
                            st.warning("The selected file did not contain any data rows to import.")
                        else:
                            existing_imports = st.session_state.get("pipeline_import_rows")
                            if isinstance(existing_imports, pd.DataFrame) and not existing_imports.empty:
                                st.session_state["pipeline_import_rows"] = pd.concat([existing_imports, imported_rows], ignore_index=True)
                            else:
                                st.session_state["pipeline_import_rows"] = imported_rows
                            st.success(f"Imported {len(imported_rows)} rows into the Pipeline grid.")
                            st.caption("Matched columns: " + ", ".join(matched_columns))
                            if unmapped_columns:
                                st.warning("Ignored columns: " + ", ".join(unmapped_columns))
                            st.rerun()
                    except Exception as exc:
                        st.error("Could not import the selected file.")
                        st.exception(exc)
            with clear_import_col:
                if st.button("Clear Imported Rows", use_container_width=True):
                    st.session_state["pipeline_import_rows"] = None
                    st.rerun()

            imported_rows_state = st.session_state.get("pipeline_import_rows")
            if isinstance(imported_rows_state, pd.DataFrame) and not imported_rows_state.empty:
                st.info(f"Imported rows waiting in Pipeline: {len(imported_rows_state)}")

        add_row_left, reset_width_left, _ = st.columns([1, 1.2, 3.8])
        with add_row_left:
            if st.button("Add Pipeline Row", use_container_width=True):
                st.session_state["pipeline_new_rows_count"] = int(st.session_state.get("pipeline_new_rows_count", 0)) + 1
                st.rerun()
        with reset_width_left:
            if st.button("Reset Column Widths", use_container_width=True):
                st.session_state.pop("pipeline_columns_state", None)
                set_persistent_columns_state("Pipeline", None)
                st.rerun()

        updated_pipeline = render_wide_table(
            pipeline_input_df,
            "Pipeline",
            height=520,
            allow_data_entry=True,
            allow_title_edit=is_manager,
        )
        st.session_state["pipeline_updates"] = updated_pipeline

        with save_button_slot.container():
            _, button_col = st.columns([3.2, 1.8])
            with button_col:
                if st.button("Save Pipeline Changes to Workbook", type="primary", use_container_width=True):
                    try:
                        updated_cells = save_activity_changes(
                            workbook_path,
                            pipeline_updates=st.session_state.get("pipeline_updates"),
                        )
                        load_data.clear()
                        st.session_state["pipeline_new_rows_count"] = 0
                        st.session_state["pipeline_import_rows"] = None
                        st.success(f"Saved pipeline updates to workbook ({updated_cells} cells updated).")
                        st.rerun()
                    except Exception as exc:
                        st.error("Could not save pipeline changes to workbook.")
                        st.exception(exc)

        st.download_button(
            label="Download filtered pipeline CSV",
            data=updated_pipeline.to_csv(index=False).encode("utf-8"),
            file_name="sales_pipeline_filtered.csv",
            mime="text/csv",
            use_container_width=True,
        )
    elif selected_page == "Blog":
        render_blog_page(source_sales, is_manager)
    elif selected_page == "Team":
        section_banner("Team", "Summary report view only. Data entry is disabled on this page.")
        team_reset_col, _ = st.columns([1.3, 4.7])
        with team_reset_col:
            if st.button("Reset Team Column Widths", use_container_width=True):
                st.session_state.pop("team_columns_state", None)
                st.session_state.pop("team_columns_state_sig", None)
                set_persistent_columns_state("Team", None)
                st.rerun()
        render_wide_table(team, "Team", height=520, allow_data_entry=False, allow_title_edit=False)
    elif selected_page == "COI":
        section_banner("COI", "Keep referral development records separate and easy to reach from the main navigation.")
        coi_save_slot = st.empty()
        if "coi_new_rows_count" not in st.session_state:
            st.session_state["coi_new_rows_count"] = 0

        coi_input_df = append_blank_rows(coi.copy(), int(st.session_state.get("coi_new_rows_count", 0)))
        coi_input_df = apply_new_row_defaults(coi_input_df, "COI")
        add_coi_left, reset_coi_left, _ = st.columns([1, 1.6, 3.4])
        with add_coi_left:
            if st.button("Add COI Row", use_container_width=True):
                st.session_state["coi_new_rows_count"] = int(st.session_state.get("coi_new_rows_count", 0)) + 1
                st.rerun()
        with reset_coi_left:
            if st.button("Reset COI Column Widths", use_container_width=True):
                st.session_state.pop("coi_columns_state", None)
                st.session_state.pop("coi_columns_state_sig", None)
                set_persistent_columns_state("COI", None)
                st.rerun()

        updated_coi = render_wide_table(coi_input_df, "COI", height=520, allow_data_entry=True, allow_title_edit=is_manager)
        st.session_state["coi_updates"] = updated_coi

        with coi_save_slot.container():
            _, coi_button_col = st.columns([3.2, 1.8])
            with coi_button_col:
                if st.button("Save COI Changes to Workbook", type="primary", use_container_width=True):
                    try:
                        updated_cells = save_activity_changes(
                            workbook_path,
                            coi_updates=st.session_state.get("coi_updates"),
                        )
                        load_data.clear()
                        st.session_state["coi_new_rows_count"] = 0
                        st.success(f"Saved COI updates to workbook ({updated_cells} cells updated).")
                        st.rerun()
                    except Exception as exc:
                        st.error("Could not save COI changes to workbook.")
                        st.exception(exc)
    elif selected_page == "Lists":
        if not is_manager:
            st.warning("This page is restricted to Firm Manager access. Please sign in using the sidebar.")
        else:
            render_lists_page(workbook_path)


def main() -> None:
    st.set_page_config(
        page_title="Sales Tracker App",
        page_icon="ðŸ“ˆ",
        layout="wide",
    )
    apply_theme()
    initialize_navigation()
    app_header()

    workbook_input = str(DEFAULT_WORKBOOK)
    active_workbook = workbook_input
    default_workbook_ok, default_workbook_error = validate_workbook_path(workbook_input)
    if not default_workbook_ok:
        st.error(default_workbook_error or "Default workbook path is invalid.")
        st.stop()
    try:
        data = load_data(workbook_input)
    except Exception as exc:
        st.error("Could not load the default workbook. Update the path in the sidebar.")
        st.exception(exc)
        st.stop()

    sales_df = data["sales"]

    if "dropdown_options" not in st.session_state:
        reset_runtime_dropdown_options(data["coi"])

    filtered_sales, workbook_override, is_manager = build_sidebar(sales_df)

    if workbook_override != workbook_input:
        override_ok, override_error = validate_workbook_path(workbook_override)
        if not override_ok:
            st.warning(override_error or "Workbook path could not be loaded. Showing default workbook instead.")
            selected_page = render_navigation(is_manager)
            render_page(selected_page, sales_df, filtered_sales, data["team"], data["coi"], data["stats"], active_workbook, is_manager)
            return
        try:
            data = load_data(workbook_override)
            sales_df = data["sales"]
            reset_runtime_dropdown_options(data["coi"])
            filtered_sales, _, is_manager = build_sidebar(sales_df)
            active_workbook = workbook_override
        except Exception as exc:
            st.warning("Workbook path could not be loaded. Showing default workbook instead.")
            st.exception(exc)

    selected_page = render_navigation(is_manager)
    render_page(selected_page, sales_df, filtered_sales, data["team"], data["coi"], data["stats"], active_workbook, is_manager)


if __name__ == "__main__":
    main()


